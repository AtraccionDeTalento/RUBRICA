"""
People Analytics - Evaluación de Talento Docente
Servidor Flask con interfaz de análisis inteligente
Talento & Cultura USIL | Version 3.0
"""
from flask import Flask, render_template, jsonify, send_from_directory, request
import os
import sys
import re
import threading
import time
import json
from datetime import datetime
import tempfile
import shutil

# Importar módulos del sistema
from extractor_cvs import procesar_todos_cvs
from extractor_web_cvs import ExtractorWebCVs, get_extraction_progress, reset_extraction_progress
from motor_evaluacion import MotorEvaluacion, TOTAL_MAXIMO
from config import RESULTADOS_DIR, EXCEL_LINKS_PATH, BASE_DIR, INFORMACION_VALIDACION_PATH, REQUERIMIENTOS_DOCENTES_PATH, LINKS_DIR, LEGAJOS_INPUT_DIR
import pandas as pd

# Configurar rutas para templates y static (funciona tanto en desarrollo como en .exe)
def get_app_paths():
      """Obtiene las rutas correctas para templates y static"""
      if getattr(sys, 'frozen', False):
             # Ejecutando como .exe (PyInstaller)
             base_path = sys._MEIPASS
      else:
             # Ejecutando como script Python
             base_path = os.path.dirname(os.path.abspath(__file__))
      
      template_folder = os.path.join(base_path, 'templates')
      static_folder = os.path.join(base_path, 'static')
      
      return template_folder, static_folder

template_folder, static_folder = get_app_paths()
app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# Configuración para subida de archivos
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB máximo
UPLOAD_FOLDER = os.path.join(tempfile.gettempdir(), 'evaluacion_docente_uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Variables globales para el estado del proceso
# Se inicializa vacío - requiere nueva evaluación cada vez que se reinicia el servidor
estado_proceso = {
      "paso_actual": 0,
      "mensaje": "Listo para evaluar. Haz clic en 'Iniciar Evaluación'.",
      "porcentaje": 0,
      "completado": False,
      "resultado": None,
      "error": None,
      # Nuevos campos para tracking en tiempo real
      "tiempo_inicio": None,
      "tiempo_estimado_restante": None,
      "cvs_procesados": 0,
      "cvs_total": 0,
      "cvs_exitosos": 0,
      "cvs_con_error": 0,
      "velocidad_cvs_por_segundo": 0,
      "registros_con_error": []
}

# Variable global para almacenar el archivo subido (ruta completa)
archivo_excel_subido = None

# Variable global para almacenar el contenido del archivo en memoria (BytesIO)
archivo_contenido_memoria = None

# Ruta fija para el archivo Excel subido por el usuario (dentro del proyecto)
ARCHIVO_SUBIDO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'archivo_usuario_subido.xlsx')

# LIMPIAR archivo subido de sesiones anteriores al iniciar
# Esto asegura que siempre se use LINKS como fuente predeterminada a menos que se suba un archivo nuevo
if os.path.exists(ARCHIVO_SUBIDO_PATH):
       try:
               os.remove(ARCHIVO_SUBIDO_PATH)
               print("[OK] Archivo de sesion anterior eliminado: " + ARCHIVO_SUBIDO_PATH)
               print("   -> Se usara la carpeta LINKS como fuente predeterminada")
       except Exception as e:
               print("[WARN] No se pudo eliminar archivo anterior: " + str(e))

# Variable global para almacenar la carpeta de PDFs subida
carpeta_pdfs_subida = None

# Modo rapido: procesa menos datos (omite extraccion CTI) para acelerar evaluaciones.
modo_rapido_activo = False

# Links de CTI Vitae agregados manualmente desde el panel (sin pasar por Excel).
links_manuales_cti = []

# Historial persistente por persona para recuperacion rapida.
HISTORIAL_PERSONAS_PATH = os.path.join(RESULTADOS_DIR, 'historial_personas.json')

# Controles de automatizacion de perfiles experimentales.
# El usuario puede resolver cada caso manualmente con botones (Legajos/Web) desde la UI.
AUTO_INDEXACION_LEGAJOS = False
AUTO_RECONSTRUCCION_WEB_EXPERIMENTAL = False

def cargar_datos_requerimientos_para_match() -> dict:
      """
      Carga los datos del archivo de requerimientos para hacer match por nombre/DNI.
      Retorna un diccionario con la información indexada por DNI y por nombre.
      """
      datos_match = {
             "por_dni": {},        # {dni_str_8dig: {facultad, carrera, nombre}}
             "por_nombre": {},     # {nombre_upper_limpio: {facultad, carrera, dni}}
             "por_palabras": {},   # {frozenset(palabras): {facultad, carrera}} — tolera orden inverso
      }
      
      try:
             # Buscar el archivo en varias ubicaciones posibles
             ruta_req = None
             candidatos_ruta = [
                    REQUERIMIENTOS_DOCENTES_PATH,
                    os.path.join(os.path.dirname(__file__), 'archivo_usuario_subido.xlsx'),
                    os.path.join(os.path.expanduser('~'), 'Downloads', 'Requerimiento docentes 2026-1.xlsx'),
             ]
             for ruta in candidatos_ruta:
                    if os.path.exists(ruta):
                           ruta_req = ruta
                           break
             if ruta_req is None:
                    print("   [!] No se encontro archivo de requerimientos para match")
                    return datos_match

             # La hoja '2026.1' tiene una fila de título arriba, así que la fila de
             # encabezados real puede estar en 0 o en 1. Probamos ambas y nos quedamos
             # con la que tenga columnas FACULTAD/CANDIDATO reconocibles.
             df = None
             for header_row in (0, 1):
                    tmp = pd.read_excel(ruta_req, sheet_name='2026.1', header=header_row)
                    cols_upper = [str(c).upper().strip() for c in tmp.columns]
                    if 'FACULTAD' in cols_upper and any(c == 'CANDIDATO' or 'APELLIDOS' in c for c in cols_upper):
                           df = tmp
                           break
             if df is None:
                    print("   [!] No se reconocieron columnas en la hoja '2026.1' del archivo de requerimientos")
                    return datos_match

             # Detectar nombres reales de columnas (admite 'CANDIDATO' o 'APELLIDOS Y NOMBRES...')
             col_nombre = col_dni = col_facultad = col_carrera = None
             for c in df.columns:
                    cu = str(c).upper().strip()
                    if col_nombre is None and (cu == 'CANDIDATO' or 'APELLIDOS' in cu):
                           col_nombre = c
                    if col_dni is None and cu == 'DNI':
                           col_dni = c
                    if col_facultad is None and cu == 'FACULTAD':
                           col_facultad = c
                    if col_carrera is None and cu == 'CARRERA':
                           col_carrera = c

             for idx, row in df.iterrows():
                    nombre_raw = str(row.get(col_nombre, '')).strip()
                    dni_raw = str(row.get(col_dni, '')).strip()
                    facultad = str(row.get(col_facultad, '')).strip()
                    carrera = str(row.get(col_carrera, '')).strip()
                    
                    # Limpiar valores
                    if nombre_raw.lower() in ['nan', 'none', '']:
                           nombre_raw = ''
                    if dni_raw.lower() in ['nan', 'none', '']:
                           dni_raw = ''
                    if facultad.lower() in ['nan', 'none', '']:
                           facultad = ''
                    if carrera.lower() in ['nan', 'none', '']:
                           carrera = ''
                    
                    # Normalizar DNI: si viene como float (16756158.0) convertir a int primero
                    try:
                           dni_norm = str(int(float(dni_raw))).strip()
                    except (ValueError, TypeError):
                           dni_norm = dni_raw.replace(' ', '').replace('.', '').replace('-', '').upper()

                    # Normalizar nombre (mayúsculas, sin espacios extra, sin comas)
                    nombre_norm = ' '.join(nombre_raw.upper().replace(',', ' ').split())
                    # Conjunto de palabras (tolera orden APELLIDO, NOMBRE vs NOMBRE APELLIDO)
                    palabras_key = frozenset(w for w in nombre_raw.upper().replace(',', ' ').split() if len(w) > 1)

                    payload = {"facultad": facultad, "carrera": carrera, "nombre": nombre_raw, "dni": dni_norm}

                    if dni_norm and facultad:
                           datos_match["por_dni"][dni_norm] = payload

                    if nombre_norm and facultad:
                           datos_match["por_nombre"][nombre_norm] = payload

                    if palabras_key and len(palabras_key) >= 2 and facultad:
                           datos_match["por_palabras"][palabras_key] = payload
             
             pass  # log removed (encoding issue)
             
      except Exception as e:
             print(f"   [!] Error cargando datos para match: {e}")
      
      return datos_match


def _url_persona(e: dict) -> str:
       """Devuelve el link CTI de la persona buscándolo en los campos donde puede
       quedar guardado (url / link_cti / archivo). 'archivo' a veces contiene texto
       que no es una URL (p.ej. "5 PDFs del legajo"), por eso se valida que sea http."""
       for clave in ('url', 'link_cti', 'archivo'):
              valor = str(e.get(clave, '') or '').strip()
              if valor.startswith('http'):
                     return valor
       return ''


def _clave_persona(dni: str, nombre: str) -> str:
       dni_n = str(dni or '').strip()
       if dni_n:
              return f"dni:{dni_n}"
       return f"nom:{_normalizar_para_match(nombre or '')}"


def _guardar_historial_personas(evaluaciones: list) -> None:
       """Persiste historico por persona manteniendo las ultimas 20 entradas."""
       try:
              if os.path.exists(HISTORIAL_PERSONAS_PATH):
                     with open(HISTORIAL_PERSONAS_PATH, 'r', encoding='utf-8') as f:
                            historial = json.load(f)
              else:
                     historial = {}

              ahora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
              for e in evaluaciones:
                     clave = _clave_persona(e.get('dni', ''), e.get('nombre', ''))
                     if not clave or clave.endswith('nom:'):
                            continue

                     registro = {
                            'fecha': ahora,
                            'nombre': e.get('nombre', ''),
                            'dni': e.get('dni', ''),
                            'facultad': e.get('facultad', ''),
                            'carrera': e.get('carrera', ''),
                            'total': e.get('total', 0),
                            'porcentaje': e.get('porcentaje', 0),
                            'clasificacion': e.get('clasificacion', ''),
                            'es_elegible': e.get('es_elegible', False),
                            'fuente_datos': e.get('fuente_datos', e.get('fuente', '')),
                     }

                     historial.setdefault(clave, [])
                     historial[clave].append(registro)
                     historial[clave] = historial[clave][-20:]

              with open(HISTORIAL_PERSONAS_PATH, 'w', encoding='utf-8') as f:
                     json.dump(historial, f, indent=2, ensure_ascii=False)
       except Exception as e:
              print(f"[WARN] No se pudo guardar historial por persona: {e}")


def leer_datos_desde_links_usil() -> tuple:
      """
      Lee datos desde los archivos de la carpeta LINKS de USIL.
      - Archivo principal: INFORMACION DE VALIDACION.xlsx (links CTI en columna K)
      - Archivo secundario: Requerimiento docentes 2026-1 300126.xlsx (para match de facultades)
      
      Returns:
             tuple: (urls, datos_excel, resumen_problemas)
      """
      urls = []
      datos_excel = []
      resumen_problemas = {
             "links_invalidos": [],
             "datos_faltantes": [],
             "resumen_facultades": {},
             "resumen_carreras": {}
      }
      
      try:
             print(f"\n[DIR] Leyendo datos desde carpeta LINKS...")
             
             # Verificar que existen los archivos
             if not os.path.exists(INFORMACION_VALIDACION_PATH):
                    pass  # log removed (encoding issue)
                    return urls, datos_excel, resumen_problemas
             
             # Cargar datos del archivo de requerimientos para hacer match
             pass  # log removed (encoding issue)
             datos_match = cargar_datos_requerimientos_para_match()
             
             # Leer archivo principal: INFORMACION DE VALIDACION.xlsx
             print(f"   [PDF] Leyendo: INFORMACION DE VALIDACION.xlsx")
             df = pd.read_excel(INFORMACION_VALIDACION_PATH, sheet_name=0, header=0)
             
             print(f"   Total registros: {len(df)}")
             print(f"   Columnas: {list(df.columns)[:12]}")
             
             # Mapear columnas (basado en análisis previo)
             # A: N°, B: FACULTAD, C: CARRERA, I: CANDIDATO, J: DNI, K: CTI
             col_facultad = 'FACULTAD'
             col_carrera = 'CARRERA'
             col_nombre = 'CANDIDATO'
             col_dni = 'DNI'
             col_cti = 'CTI'
             
             # Verificar que las columnas existen
             cols_disponibles = list(df.columns)
             if col_nombre not in cols_disponibles:
                    # Buscar alternativa
                    for c in cols_disponibles:
                           if 'CANDIDATO' in str(c).upper():
                                  col_nombre = c
                                  break
             
             print(f"   Usando columnas: NOMBRE={col_nombre}, DNI={col_dni}, CTI={col_cti}, FACULTAD={col_facultad}")
             
             for idx, row in df.iterrows():
                    # Extraer datos del archivo principal
                    nombre_raw = str(row.get(col_nombre, '')).strip()
                    dni_raw = str(row.get(col_dni, '')).strip()
                    cti_raw = str(row.get(col_cti, '')).strip()
                    facultad_raw = str(row.get(col_facultad, '')).strip()
                    carrera_raw = str(row.get(col_carrera, '')).strip()
                    
                    # Limpiar valores
                    nombre = nombre_raw if nombre_raw.lower() not in ['nan', 'none', ''] else ''
                    dni = dni_raw.replace(' ', '') if dni_raw.lower() not in ['nan', 'none', ''] else ''
                    cti = cti_raw if cti_raw.lower() not in ['nan', 'none', ''] else ''
                    facultad = facultad_raw if facultad_raw.lower() not in ['nan', 'none', ''] else ''
                    carrera = carrera_raw if carrera_raw.lower() not in ['nan', 'none', ''] else ''
                    
                    # Si no hay facultad, buscar en el archivo de requerimientos por DNI o nombre
                    if not facultad:
                           # Primero intentar por DNI
                           dni_norm = dni.replace('.', '').replace('-', '').upper()
                           if dni_norm and dni_norm in datos_match["por_dni"]:
                                  match_data = datos_match["por_dni"][dni_norm]
                                  facultad = match_data.get("facultad", "")
                                  carrera = match_data.get("carrera", "") if not carrera else carrera
                           
                           # Si no encontró, intentar por nombre
                           if not facultad and nombre:
                                  nombre_norm = ' '.join(nombre.upper().split())
                                  if nombre_norm in datos_match["por_nombre"]:
                                         match_data = datos_match["por_nombre"][nombre_norm]
                                         facultad = match_data.get("facultad", "")
                                         carrera = match_data.get("carrera", "") if not carrera else carrera
                    
                    # Valor por defecto si no se encuentra facultad
                    if not facultad:
                           facultad = 'SIN FACULTAD ASIGNADA'
                    if not carrera:
                           carrera = 'SIN CARRERA'
                    
                    # Inicializar contadores por facultad
                    if facultad not in resumen_problemas["resumen_facultades"]:
                           resumen_problemas["resumen_facultades"][facultad] = {
                                  "total": 0,
                                  "con_link_valido": 0,
                                  "link_invalido": 0,
                                  "sin_link": 0,
                                  "datos_incompletos": 0,
                                  "personas_link_invalido": [],
                                  "personas_sin_link": [],
                                  "personas_datos_faltantes": [],
                                  "carreras": {}
                           }
                    
                    if carrera not in resumen_problemas["resumen_facultades"][facultad]["carreras"]:
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera] = {
                                  "total": 0,
                                  "con_link_valido": 0,
                                  "sin_link": 0,
                                  "link_invalido": 0,
                                  "personas": []
                           }
                    
                    resumen_problemas["resumen_facultades"][facultad]["total"] += 1
                    resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["total"] += 1
                    
                    # Verificar estado del link CTI
                    cti_lower = cti.lower()
                    es_link_invalido = False
                    motivo_invalido = ""
                    
                    if not cti:
                           # Sin link
                           resumen_problemas["resumen_facultades"][facultad]["sin_link"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["sin_link"] += 1
                           persona_data = {
                                  "nombre": nombre if nombre else "SIN NOMBRE",
                                  "dni": dni,
                                  "carrera": carrera,
                                  "motivo": "SIN LINK CTI"
                           }
                           resumen_problemas["resumen_facultades"][facultad]["personas_sin_link"].append(persona_data)
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["personas"].append(persona_data)
                           resumen_problemas["links_invalidos"].append({
                                  "nombre": nombre if nombre else "SIN NOMBRE",
                                  "dni": dni,
                                  "facultad": facultad,
                                  "carrera": carrera,
                                  "motivo": "SIN LINK CTI"
                           })
                           # Sin link CTI â†’ guardar para búsqueda automática en legajos
                           if nombre or dni:
                                  resumen_problemas.setdefault("datos_sin_link", []).append({
                                         "nombre": nombre, "dni": dni, "facultad": facultad, "carrera": carrera
                                  })
                    elif 'en construccion' in cti_lower or 'en construcción' in cti_lower:
                           es_link_invalido = True
                           motivo_invalido = "EN CONSTRUCCIÓN"
                    elif 'no tiene' in cti_lower:
                           es_link_invalido = True
                           motivo_invalido = "NO TIENE CTI"
                    elif not cti.startswith('http') or 'ctivitae.concytec.gob.pe' not in cti:
                           es_link_invalido = True
                           motivo_invalido = "LINK NO VÁLIDO"
                    else:
                           # Link válido
                           resumen_problemas["resumen_facultades"][facultad]["con_link_valido"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["con_link_valido"] += 1
                           urls.append(cti)
                           datos_excel.append({
                                  'url': cti,
                                  'dni': dni,
                                  'nombre': nombre,
                                  'facultad': facultad,
                                  'carrera': carrera
                           })
                    
                    if es_link_invalido:
                           resumen_problemas["resumen_facultades"][facultad]["link_invalido"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["link_invalido"] += 1
                           persona_data = {
                                  "nombre": nombre if nombre else "SIN NOMBRE",
                                  "dni": dni,
                                  "carrera": carrera,
                                  "motivo": motivo_invalido
                           }
                           resumen_problemas["resumen_facultades"][facultad]["personas_link_invalido"].append(persona_data)
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["personas"].append(persona_data)
                           resumen_problemas["links_invalidos"].append({
                                  "nombre": nombre if nombre else "SIN NOMBRE",
                                  "dni": dni,
                                  "facultad": facultad,
                                  "carrera": carrera,
                                  "motivo": motivo_invalido
                           })
                           # Link inválido/EN CONSTRUCCIÓN â†’ guardar para búsqueda automática en legajos
                           if nombre or dni:
                                  resumen_problemas.setdefault("datos_sin_link", []).append({
                                         "nombre": nombre, "dni": dni, "facultad": facultad, "carrera": carrera
                                  })
             
             print(f"\n[OK] Lectura completada:")
             pass  # log removed (encoding issue)
             print(f"   Con DNI: {sum(1 for d in datos_excel if d['dni'])}")
             print(f"   Con nombre: {sum(1 for d in datos_excel if d['nombre'])}")
             pass  # log removed (encoding issue)
             
      except Exception as e:
             print(f"[ERR] Error leyendo archivos de LINKS: {e}")
             import traceback
             traceback.print_exc()
      
      return urls, datos_excel, resumen_problemas


def leer_datos_desde_archivo_subido(ruta_archivo: str) -> tuple:
      """
      Lee URLs, DNIs y nombres desde un archivo Excel subido por el usuario.
      Busca automáticamente columnas con URLs de CTI Vitae.
      También genera resumen de problemas (links faltantes/inválidos).
      
      Returns:
             tuple: (urls, datos_excel, resumen_problemas)
      """
      from io import BytesIO
      
      urls = []
      datos_excel = []
      resumen_problemas = {
             "links_invalidos": [],
             "datos_faltantes": [],
             "resumen_facultades": {},
             "resumen_carreras": {}
      }
      
      try:
             print(f"\n[DIR] Leyendo datos desde archivo subido: {os.path.basename(ruta_archivo)}")
             
             # Leer el archivo completo en memoria para evitar [WinError 32]
             # Esto evita conflictos cuando el archivo está siendo usado por otro proceso
             with open(ruta_archivo, 'rb') as f:
                    contenido_archivo = BytesIO(f.read())
             pass  # log removed (encoding issue)
             
             # Leer todas las hojas disponibles desde memoria
             xl = pd.ExcelFile(contenido_archivo)
             hojas_disponibles = xl.sheet_names
             print(f"   Hojas disponibles: {hojas_disponibles}")
             
             # Función para verificar si tiene las columnas requeridas
             def tiene_columnas_requeridas(df_check):
                    """Verifica si el DataFrame tiene las columnas clave: CANDIDATO, DNI, CTI"""
                    cols_upper = [str(c).upper().strip() for c in df_check.columns]
                    tiene_candidato = any(c == 'CANDIDATO' or 'APELLIDOS' in c for c in cols_upper)
                    tiene_dni = any(c == 'DNI' for c in cols_upper)
                    tiene_cti = any(c == 'CTI' or c == 'LINK CTI' or 'CTI (LINKS)' in c for c in cols_upper)
                    tiene_urls = False
                    for col in df_check.columns:
                           sample = df_check[col].dropna().astype(str).head(5)
                           if sample.str.contains('ctivitae.concytec.gob.pe', case=False).any():
                                  tiene_urls = True
                                  break
                    return (tiene_candidato and tiene_dni) or (tiene_dni and tiene_cti) or tiene_urls or tiene_cti
             
             def contar_urls_validas(df_check):
                    """Cuenta cuántas URLs válidas de CTI Vitae hay en el DataFrame"""
                    count = 0
                    for col in df_check.columns:
                           for val in df_check[col].dropna():
                                  val_str = str(val).strip()
                                  if 'ctivitae.concytec.gob.pe' in val_str and val_str.startswith('http'):
                                         count += 1
                    return count
             
             # Priorizar hojas con nombre que contenga año (ej: "2026.1", "2025.1")
             hojas_prioritarias = [h for h in hojas_disponibles if any(str(y) in h for y in range(2020, 2030))]
             hojas_restantes = [h for h in hojas_disponibles if h not in hojas_prioritarias]
             hojas_ordenadas = hojas_prioritarias + hojas_restantes
             
             df = None
             hoja_usada = None
             
             # Intentar diferentes configuraciones de header - COMPARAR URLs
             for hoja in hojas_ordenadas:
                    try:
                           # Probar AMBOS headers y contar URLs
                           contenido_archivo.seek(0)
                           df_h0 = pd.read_excel(contenido_archivo, sheet_name=hoja, header=0)
                           
                           contenido_archivo.seek(0)
                           df_h1 = pd.read_excel(contenido_archivo, sheet_name=hoja, header=1)
                           
                           # Contar URLs en cada caso
                           urls_h0 = contar_urls_validas(df_h0)
                           urls_h1 = contar_urls_validas(df_h1)
                           
                           print(f"   Hoja: h0={urls_h0} h1={urls_h1} URLs")

                           # Calcular cuántas columnas clave reconoce cada opción.
                           # FACULTAD y CARRERA suman 2 puntos cada una porque son críticas para
                           # los filtros; las demás suman 1.  Si header=1 tiene mejor puntaje,
                           # tiene prioridad AUNQUE el conteo de URLs sea igual (el empate es la
                           # situación habitual en la hoja 2026.1, que tiene fila de título arriba).
                           def _score_cols(df_check):
                                  cu = [str(c).upper().strip() for c in df_check.columns]
                                  s = 0
                                  if 'FACULTAD' in cu: s += 2
                                  if 'CARRERA' in cu: s += 2
                                  if any(c == 'CANDIDATO' or 'APELLIDOS' in c for c in cu): s += 1
                                  if 'DNI' in cu: s += 1
                                  if any(c in ('CTI', 'LINK CTI') or 'CTI (LINKS)' in c for c in cu): s += 1
                                  return s

                           score_h0 = _score_cols(df_h0)
                           score_h1 = _score_cols(df_h1)
                           print(f"   Score columnas: header=0={score_h0}, header=1={score_h1}")

                           # Preferir el header con mejor cobertura de columnas.
                           # Usar URLs solo como desempate final.
                           if score_h1 > score_h0 and len(df_h1) > 0 and tiene_columnas_requeridas(df_h1):
                                  df = df_h1
                                  hoja_usada = hoja
                                  pass  # log removed (encoding issue)
                                  break
                           elif score_h0 >= score_h1 and len(df_h0) > 0 and tiene_columnas_requeridas(df_h0) and urls_h0 >= urls_h1:
                                  df = df_h0
                                  hoja_usada = hoja
                                  pass  # log removed (encoding issue)
                                  break
                           elif len(df_h1) > 0 and tiene_columnas_requeridas(df_h1):
                                  df = df_h1
                                  hoja_usada = hoja
                                  pass  # log removed (encoding issue)
                                  break
                           elif len(df_h0) > 0 and urls_h0 > 0:
                                  # Fallback: usar header=0 si tiene alguna URL
                                  df = df_h0
                                  hoja_usada = hoja
                                  pass  # log removed (encoding issue)
                                  break
                                  
                    except Exception as e:
                           print(f"   [!] Error en hoja {hoja}: {e}")
                           continue
             
             if df is None:
                    print("[!] No se pudo leer ninguna hoja del Excel con columnas requeridas (CANDIDATO, DNI, CTI)")
                    return urls, datos_excel, resumen_problemas
             
             print(f"   Usando hoja: {hoja_usada}")
             print(f"   Total filas: {len(df)}")
             print(f"   Columnas: {list(df.columns)}")
             
             # Buscar columnas con URLs CTI, DNI, nombres, facultad y carrera
             col_url = None
             col_dni = None
             col_nombre = None
             col_facultad = None
             col_carrera = None
             col_tipo_candidato = None
             col_puesto = None
             col_cat_practitioner = None

             for col in df.columns:
                    col_str = str(col).upper().strip()
                    
                    # Buscar columna de URL (CTI)
                    if col_str == 'CTI' or 'VITAE' in col_str:
                           col_url = col
                    elif col_url is None and any(x in col_str for x in ['URL', 'LINK', 'ENLACE']):
                           col_url = col
                    
                    # Buscar columna de DNI
                    if col_str == 'DNI' or 'DOCUMENTO' in col_str:
                           col_dni = col
                    
                    # Buscar columna de nombre/candidato
                    # Priorizar "CANDIDATO" exacto, luego "APELLIDOS Y NOMBRES"
                    if col_str == 'CANDIDATO':
                           col_nombre = col
                    elif col_nombre is None and ('APELLIDO' in col_str or ('NOMBRE' in col_str and 'CANDIDATO' in col_str)):
                           col_nombre = col
                    elif col_nombre is None and col_str == 'NOMBRE' or col_str == 'NOMBRES':
                           col_nombre = col
                    
                    # Buscar columna de facultad (varios nombres posibles)
                    if col_str == 'FACULTAD':
                           col_facultad = col
                    elif col_facultad is None and 'FACULTAD' in col_str:
                           col_facultad = col
                    elif col_facultad is None and 'UNIDAD' in col_str and 'ACAD' in col_str:
                           col_facultad = col
                    elif col_facultad is None and col_str in ('UNIDAD ACADEMICA', 'UNIDAD ACADÉMICA'):
                           col_facultad = col

                    # Buscar columna de carrera
                    if col_str == 'CARRERA' or 'PROGRAMA' in col_str:
                           col_carrera = col
                    elif col_carrera is None and 'ESCUELA' in col_str:
                           col_carrera = col

                    # Buscar columna de TIPO DE CANDIDATO (NUEVO, REEMPLAZO, PASA TP A TC...)
                    if col_tipo_candidato is None and 'TIPO' in col_str and 'CANDIDATO' in col_str:
                           col_tipo_candidato = col

                    # Buscar columna de PUESTO (Docente TC / TP), evitando "TIPO PUESTO"
                    if col_str == 'PUESTO':
                           col_puesto = col
                    elif col_puesto is None and 'PUESTO' in col_str and 'TIPO' not in col_str:
                           col_puesto = col

                    # Buscar columna de CATEGORIA PRACTITIONER
                    if col_cat_practitioner is None and 'PRACTITIONER' in col_str:
                           col_cat_practitioner = col

             # Si no encontró columna de URL, buscar en los datos
             if col_url is None:
                    for col in df.columns:
                           sample_data = df[col].dropna().astype(str).head(10)
                           if sample_data.str.contains('ctivitae.concytec.gob.pe').any():
                                  col_url = col
                                  break
             
             if col_url is None:
                    pass  # log removed (encoding issue)
                    return urls, datos_excel, resumen_problemas
             
             pass  # log removed (encoding issue)
             print(f"        - URLs: {col_url}")
             print(f"        - DNI: {col_dni if col_dni else 'No encontrada'}")
             print(f"        - Nombre: {col_nombre if col_nombre else 'No encontrada'}")
             print(f"        - Facultad: {col_facultad if col_facultad else 'No encontrada'}")
             print(f"        - Carrera: {col_carrera if col_carrera else 'No encontrada'}")
             
             pass  # log removed (encoding issue)
             datos_match = cargar_datos_requerimientos_para_match()
             
             # Procesar filas
             for idx, row in df.iterrows():
                    url_raw = str(row.get(col_url, '')).strip()
                    dni_raw = str(row.get(col_dni, '')).strip() if col_dni else ''
                    nombre_raw = str(row.get(col_nombre, '')).strip() if col_nombre else ''
                    facultad_raw = str(row.get(col_facultad, '')).strip() if col_facultad else 'ARCHIVO SUBIDO'
                    carrera_raw = str(row.get(col_carrera, '')).strip() if col_carrera else 'SIN CARRERA'

                    # Datos de la solicitud (hoja 2026.1): tipo de candidato, puesto y categoria practitioner
                    def _limpiar_celda(valor):
                           s = str(valor).strip()
                           # \xa0 es un espacio no-rompible que aparece en celdas "vacias" de Excel
                           s = s.replace('\xa0', ' ').strip()
                           return '' if s.lower() in ['nan', 'none', ''] else s
                    tipo_candidato = _limpiar_celda(row.get(col_tipo_candidato, '')) if col_tipo_candidato else ''
                    puesto = _limpiar_celda(row.get(col_puesto, '')) if col_puesto else ''
                    categoria_practitioner = _limpiar_celda(row.get(col_cat_practitioner, '')) if col_cat_practitioner else ''

                    # Limpiar DNI
                    if dni_raw and dni_raw.lower() not in ['nan', 'none', '']:
                           dni = dni_raw.replace(' ', '')
                    else:
                           dni = ''
                    
                    # Limpiar nombre
                    if nombre_raw and nombre_raw.lower() not in ['nan', 'none', '']:
                           nombre = nombre_raw
                    else:
                           nombre = ''
                    
                    # Limpiar facultad
                    if facultad_raw and facultad_raw.lower() not in ['nan', 'none', '']:
                           facultad = facultad_raw
                    else:
                           facultad = ''
                    
                    # Limpiar carrera
                    if carrera_raw and carrera_raw.lower() not in ['nan', 'none', '']:
                           carrera = carrera_raw
                    else:
                           carrera = ''
                    
                    # Si no hay facultad, buscar en el archivo de requerimientos por DNI o nombre
                    if not facultad:
                           # Primero intentar por DNI
                           dni_norm = dni.replace('.', '').replace('-', '').upper()
                           if dni_norm and dni_norm in datos_match["por_dni"]:
                                  match_data = datos_match["por_dni"][dni_norm]
                                  facultad = match_data.get("facultad", "")
                                  carrera = match_data.get("carrera", "") if not carrera else carrera
                           
                           # Si no encontró, intentar por nombre
                           if not facultad and nombre:
                                  nombre_norm = ' '.join(nombre.upper().split())
                                  if nombre_norm in datos_match["por_nombre"]:
                                         match_data = datos_match["por_nombre"][nombre_norm]
                                         facultad = match_data.get("facultad", "")
                                         carrera = match_data.get("carrera", "") if not carrera else carrera
                    
                    if not facultad:
                           facultad = 'ARCHIVO SUBIDO'
                    if not carrera:
                           carrera = 'SIN CARRERA'
                    
                    # Inicializar contador de facultad si no existe
                    if facultad not in resumen_problemas["resumen_facultades"]:
                           resumen_problemas["resumen_facultades"][facultad] = {
                                  "total": 0,
                                  "con_link_valido": 0,
                                  "link_invalido": 0,
                                  "sin_link": 0,
                                  "datos_incompletos": 0,
                                  "personas_link_invalido": [],
                                  "personas_sin_link": [],
                                  "personas_datos_faltantes": [],
                                  "carreras": {}
                           }
                    
                    # Inicializar carrera dentro de facultad
                    if carrera not in resumen_problemas["resumen_facultades"][facultad]["carreras"]:
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera] = {
                                  "total": 0,
                                  "con_link_valido": 0,
                                  "sin_link": 0,
                                  "link_invalido": 0,
                                  "personas": []
                           }
                    
                    resumen_problemas["resumen_facultades"][facultad]["total"] += 1
                    resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["total"] += 1
                    
                    # Verificar estado del link CTI
                    url_lower = url_raw.lower()
                    es_link_invalido = False
                    motivo_invalido = ""
                    
                    if not url_raw or url_lower in ['nan', 'none', '']:
                           # Sin link
                           resumen_problemas["resumen_facultades"][facultad]["sin_link"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["sin_link"] += 1
                           persona_data = {
                                  "nombre": nombre if nombre else "SIN NOMBRE",
                                  "dni": dni,
                                  "carrera": carrera,
                                  "motivo": "SIN LINK CTI"
                           }
                           resumen_problemas["resumen_facultades"][facultad]["personas_sin_link"].append(persona_data)
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["personas"].append(persona_data)
                           resumen_problemas["links_invalidos"].append({
                                  "nombre": nombre if nombre else "SIN NOMBRE",
                                  "dni": dni,
                                  "facultad": facultad,
                                  "carrera": carrera,
                                  "motivo": "SIN LINK CTI"
                           })
                           # Sin link CTI â†’ guardar para búsqueda automática en legajos
                           if nombre or dni:
                                  resumen_problemas.setdefault("datos_sin_link", []).append({
                                         "nombre": nombre, "dni": dni, "facultad": facultad, "carrera": carrera
                                  })
                    elif 'en construccion' in url_lower or 'en construcción' in url_lower:
                           es_link_invalido = True
                           motivo_invalido = "EN CONSTRUCCIÓN"
                    elif 'no tiene' in url_lower:
                           es_link_invalido = True
                           motivo_invalido = "NO TIENE CTI"
                    elif not url_raw.startswith('http') or 'ctivitae.concytec.gob.pe' not in url_raw:
                           es_link_invalido = True
                           motivo_invalido = "LINK NO VÁLIDO"
                    else:
                           # Link válido
                           resumen_problemas["resumen_facultades"][facultad]["con_link_valido"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["con_link_valido"] += 1
                           urls.append(url_raw)
                           datos_excel.append({
                                  'url': url_raw,
                                  'dni': dni,
                                  'nombre': nombre,
                                  'facultad': facultad,
                                  'carrera': carrera,
                                  'tipo_candidato': tipo_candidato,
                                  'puesto': puesto,
                                  'categoria_practitioner': categoria_practitioner
                           })
                    
                    if es_link_invalido:
                           resumen_problemas["resumen_facultades"][facultad]["link_invalido"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["link_invalido"] += 1
                           persona_data = {
                                  "nombre": nombre if nombre else "SIN NOMBRE",
                                  "dni": dni,
                                  "carrera": carrera,
                                  "motivo": motivo_invalido
                           }
                           resumen_problemas["resumen_facultades"][facultad]["personas_link_invalido"].append(persona_data)
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["personas"].append(persona_data)
                           resumen_problemas["links_invalidos"].append({
                                  "nombre": nombre if nombre else "SIN NOMBRE",
                                  "dni": dni,
                                  "facultad": facultad,
                                  "carrera": carrera,
                                  "motivo": motivo_invalido
                           })
                           # Link inválido/EN CONSTRUCCIÓN â†’ guardar para búsqueda automática en legajos
                           if nombre or dni:
                                  resumen_problemas.setdefault("datos_sin_link", []).append({
                                         "nombre": nombre, "dni": dni, "facultad": facultad, "carrera": carrera
                                  })
             
             pass  # log removed (encoding issue)
             print(f"   Con DNI: {sum(1 for d in datos_excel if d['dni'])}")
             print(f"   Con nombre: {sum(1 for d in datos_excel if d['nombre'])}")
             pass  # log removed (encoding issue)
             
      except Exception as e:
             print(f"[ERR] Error leyendo archivo subido: {e}")
             import traceback
             traceback.print_exc()
      
      return urls, datos_excel, resumen_problemas


def leer_datos_desde_memoria(contenido_archivo) -> tuple:
      """
      Lee URLs, DNIs y nombres desde un BytesIO en memoria.
      Busca automáticamente columnas con URLs de CTI Vitae.
      También genera resumen de problemas (links faltantes/inválidos).
      
      Args:
             contenido_archivo: BytesIO con el contenido del Excel
      
      Returns:
             tuple: (urls, datos_excel, resumen_problemas)
      """
      urls = []
      datos_excel = []
      resumen_problemas = {
             "links_invalidos": [],
             "datos_faltantes": [],
             "resumen_facultades": {},
             "resumen_carreras": {}
      }
      
      try:
             print(f"\n[DIR] Procesando archivo desde memoria...")
             
             # Reiniciar posición del buffer
             contenido_archivo.seek(0)
             
             # Leer todas las hojas disponibles desde memoria
             xl = pd.ExcelFile(contenido_archivo)
             hojas_disponibles = xl.sheet_names
             print(f"   Hojas disponibles: {hojas_disponibles}")
             
             # Función para verificar si tiene las columnas requeridas
             def tiene_columnas_requeridas(df_check):
                    """Verifica si el DataFrame tiene las columnas clave: CANDIDATO, DNI, CTI"""
                    cols_upper = [str(c).upper().strip() for c in df_check.columns]
                    tiene_candidato = any(c == 'CANDIDATO' or 'APELLIDOS' in c for c in cols_upper)
                    tiene_dni = any(c == 'DNI' for c in cols_upper)
                    tiene_cti = any(c == 'CTI' or c == 'LINK CTI' or 'CTI (LINKS)' in c for c in cols_upper)
                    tiene_urls = False
                    for col in df_check.columns:
                           sample = df_check[col].dropna().astype(str).head(5)
                           if sample.str.contains('ctivitae.concytec.gob.pe', case=False).any():
                                  tiene_urls = True
                                  break
                    return (tiene_candidato and tiene_dni) or (tiene_dni and tiene_cti) or tiene_urls or tiene_cti
             
             def contar_urls_validas(df_check):
                    """Cuenta cuántas URLs válidas de CTI Vitae hay en el DataFrame"""
                    count = 0
                    for col in df_check.columns:
                           for val in df_check[col].dropna():
                                  val_str = str(val).strip()
                                  if 'ctivitae.concytec.gob.pe' in val_str and val_str.startswith('http'):
                                         count += 1
                    return count
             
             # Priorizar hojas con nombre que contenga año
             hojas_prioritarias = [h for h in hojas_disponibles if any(str(y) in h for y in range(2020, 2030))]
             hojas_restantes = [h for h in hojas_disponibles if h not in hojas_prioritarias]
             hojas_ordenadas = hojas_prioritarias + hojas_restantes
             
             df = None
             hoja_usada = None
             
             for hoja in hojas_ordenadas:
                    try:
                           # IMPORTANTE: Probar header=0 PRIMERO (es más común tener encabezados en fila 1)
                           contenido_archivo.seek(0)
                           df_h0 = pd.read_excel(contenido_archivo, sheet_name=hoja, header=0)
                           
                           contenido_archivo.seek(0)
                           df_h1 = pd.read_excel(contenido_archivo, sheet_name=hoja, header=1)
                           
                           # Contar URLs en cada caso para elegir el mejor
                           urls_h0 = contar_urls_validas(df_h0)
                           urls_h1 = contar_urls_validas(df_h1)
                           
                           print(f"   Hoja '{hoja}': h0={urls_h0} URLs, h1={urls_h1} URLs")
                           
                           # Elegir el que tenga más URLs (eso indica que no se perdió ninguna como header)
                           if urls_h0 >= urls_h1 and len(df_h0) > 0 and tiene_columnas_requeridas(df_h0):
                                  df = df_h0
                                  hoja_usada = hoja
                                  pass  # log removed (encoding issue)
                                  break
                           elif len(df_h1) > 0 and tiene_columnas_requeridas(df_h1):
                                  df = df_h1
                                  hoja_usada = hoja
                                  pass  # log removed (encoding issue)
                                  break
                           elif len(df_h0) > 0 and urls_h0 > 0:
                                  # Fallback: usar header=0 si tiene alguna URL
                                  df = df_h0
                                  hoja_usada = hoja
                                  pass  # log removed (encoding issue)
                                  break
                                  
                    except Exception as e:
                           print(f"   [!] Error en hoja {hoja}: {e}")
                           continue
             
             if df is None:
                    print("[!] No se pudo leer ninguna hoja del Excel")
                    return urls, datos_excel, resumen_problemas
             
             print(f"   Usando hoja: {hoja_usada}")
             print(f"   Total filas: {len(df)}")
             print(f"   Columnas encontradas: {list(df.columns)}")
             
             # Buscar columnas - mejorado para detectar variaciones
             col_url = col_dni = col_nombre = col_facultad = col_carrera = None
             col_tipo_candidato = col_puesto = col_cat_practitioner = None
             
             for col in df.columns:
                    col_str = str(col).upper().strip()
                    
                    # Buscar columna CTI/URL (prioridad: CTI exacto, luego variaciones)
                    if col_str == 'CTI' or col_str == 'LINK CTI':
                           col_url = col
                    elif col_url is None and ('VITAE' in col_str or 'CONCYTEC' in col_str):
                           col_url = col
                    elif col_url is None and any(x in col_str for x in ['URL', 'LINK', 'ENLACE']):
                           col_url = col
                    
                    # Buscar columna DNI
                    if 'DNI' in col_str or 'DOCUMENTO' in col_str:
                           col_dni = col
                    
                    # Buscar columna de nombre/candidato
                    if 'APELLIDOS Y NOMBRES' in col_str or 'CANDIDATO' in col_str:
                           col_nombre = col
                    elif col_nombre is None and 'APELLIDO' in col_str:
                           col_nombre = col
                    elif col_nombre is None and col_str in ['NOMBRE', 'NOMBRES']:
                           col_nombre = col
                    
                    # Buscar columna de facultad
                    if 'FACULTAD' in col_str:
                           col_facultad = col
                    
                    # Buscar columna de carrera
                    if 'CARRERA' in col_str or 'PROGRAMA' in col_str:
                           col_carrera = col

                    # Buscar columna de TIPO DE CANDIDATO
                    if col_tipo_candidato is None and 'TIPO' in col_str and 'CANDIDATO' in col_str:
                           col_tipo_candidato = col

                    # Buscar columna de PUESTO
                    if col_str == 'PUESTO':
                           col_puesto = col
                    elif col_puesto is None and 'PUESTO' in col_str and 'TIPO' not in col_str:
                           col_puesto = col

                    # Buscar columna de CATEGORIA PRACTITIONER
                    if col_cat_practitioner is None and 'PRACTITIONER' in col_str:
                           col_cat_practitioner = col

             
             if col_url is None:
                    for col in df.columns:
                           sample_data = df[col].dropna().astype(str).head(10)
                           if sample_data.str.contains('ctivitae.concytec.gob.pe').any():
                                  col_url = col
                                  break
             
             if col_url is None:
                    pass  # log removed (encoding issue)
                    return urls, datos_excel, resumen_problemas
             
             pass  # log removed (encoding issue)
             print(f"        - URLs: {col_url}")
             print(f"        - DNI: {col_dni if col_dni else 'No encontrada'}")
             print(f"        - Nombre: {col_nombre if col_nombre else 'No encontrada'}")
             
             pass  # log removed (encoding issue)
             datos_match = cargar_datos_requerimientos_para_match()
             
             # Procesar filas
             for idx, row in df.iterrows():
                    url_raw = str(row.get(col_url, '')).strip()
                    dni_raw = str(row.get(col_dni, '')).strip() if col_dni else ''
                    nombre_raw = str(row.get(col_nombre, '')).strip() if col_nombre else ''
                    facultad_raw = str(row.get(col_facultad, '')).strip() if col_facultad else 'ARCHIVO SUBIDO'
                    carrera_raw = str(row.get(col_carrera, '')).strip() if col_carrera else 'SIN CARRERA'

                    def _limpiar_celda(valor):
                           s = str(valor).strip()
                           s = s.replace('\xa0', ' ').strip()
                           return '' if s.lower() in ['nan', 'none', ''] else s

                    tipo_candidato = _limpiar_celda(row.get(col_tipo_candidato, '')) if col_tipo_candidato else ''
                    puesto = _limpiar_celda(row.get(col_puesto, '')) if col_puesto else ''
                    categoria_practitioner = _limpiar_celda(row.get(col_cat_practitioner, '')) if col_cat_practitioner else ''
                    
                    dni = dni_raw.replace(' ', '') if dni_raw and dni_raw.lower() not in ['nan', 'none', ''] else ''
                    nombre = nombre_raw if nombre_raw and nombre_raw.lower() not in ['nan', 'none', ''] else ''
                    facultad = facultad_raw if facultad_raw and facultad_raw.lower() not in ['nan', 'none', ''] else ''
                    carrera = carrera_raw if carrera_raw and carrera_raw.lower() not in ['nan', 'none', ''] else ''
                    
                    # Si no hay facultad, buscar en el archivo de requerimientos por DNI o nombre
                    if not facultad:
                           # Primero intentar por DNI
                           dni_norm = dni.replace('.', '').replace('-', '').upper()
                           if dni_norm and dni_norm in datos_match["por_dni"]:
                                  match_data = datos_match["por_dni"][dni_norm]
                                  facultad = match_data.get("facultad", "")
                                  carrera = match_data.get("carrera", "") if not carrera else carrera
                           
                           # Si no encontró, intentar por nombre
                           if not facultad and nombre:
                                  nombre_norm = ' '.join(nombre.upper().split())
                                  if nombre_norm in datos_match["por_nombre"]:
                                         match_data = datos_match["por_nombre"][nombre_norm]
                                         facultad = match_data.get("facultad", "")
                                         carrera = match_data.get("carrera", "") if not carrera else carrera
                    
                    if not facultad:
                           facultad = 'ARCHIVO SUBIDO'
                    if not carrera:
                           carrera = 'SIN CARRERA'
                    
                    if facultad not in resumen_problemas["resumen_facultades"]:
                           resumen_problemas["resumen_facultades"][facultad] = {
                                  "total": 0, "sin_link": 0, "link_invalido": 0, "con_link_valido": 0,
                                  "personas_sin_link": [], "personas_link_invalido": [],
                                  "carreras": {}
                           }
                    resumen_problemas["resumen_facultades"][facultad]["total"] += 1
                    
                    if carrera not in resumen_problemas["resumen_facultades"][facultad]["carreras"]:
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera] = {
                                  "total": 0, "sin_link": 0, "link_invalido": 0, "con_link_valido": 0, "personas": []
                           }
                    resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["total"] += 1
                    
                    url_lower = url_raw.lower()
                    es_link_invalido = False
                    motivo_invalido = ""
                    
                    if not url_raw or url_lower in ['nan', 'none', '', 'sin cti', 'no tiene', '-']:
                           resumen_problemas["resumen_facultades"][facultad]["sin_link"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["sin_link"] += 1
                           persona_data = {"nombre": nombre if nombre else "SIN NOMBRE", "dni": dni, "carrera": carrera, "motivo": "SIN LINK CTI"}
                           resumen_problemas["resumen_facultades"][facultad]["personas_sin_link"].append(persona_data)
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["personas"].append(persona_data)
                           resumen_problemas["links_invalidos"].append({
                                  "nombre": nombre if nombre else "SIN NOMBRE", "dni": dni,
                                  "facultad": facultad, "carrera": carrera, "motivo": "SIN LINK CTI"
                           })
                           # Sin link CTI â†’ guardar para búsqueda automática en legajos
                           if nombre or dni:
                                  resumen_problemas.setdefault("datos_sin_link", []).append({
                                         "nombre": nombre, "dni": dni, "facultad": facultad, "carrera": carrera
                                  })
                    elif 'en construccion' in url_lower or 'en construcción' in url_lower:
                           es_link_invalido = True
                           motivo_invalido = "EN CONSTRUCCIÓN"
                    elif 'no tiene' in url_lower:
                           es_link_invalido = True
                           motivo_invalido = "NO TIENE CTI"
                    elif not url_raw.startswith('http') or 'ctivitae.concytec.gob.pe' not in url_raw:
                           es_link_invalido = True
                           motivo_invalido = "LINK NO VÁLIDO"
                    else:
                           resumen_problemas["resumen_facultades"][facultad]["con_link_valido"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["con_link_valido"] += 1
                           urls.append(url_raw)
                           datos_excel.append({
                                  'url': url_raw, 'dni': dni, 'nombre': nombre, 'facultad': facultad, 'carrera': carrera,
                                  'tipo_candidato': tipo_candidato, 'puesto': puesto, 'categoria_practitioner': categoria_practitioner
                           })
                    
                    if es_link_invalido:
                           resumen_problemas["resumen_facultades"][facultad]["link_invalido"] += 1
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["link_invalido"] += 1
                           persona_data = {"nombre": nombre if nombre else "SIN NOMBRE", "dni": dni, "carrera": carrera, "motivo": motivo_invalido}
                           resumen_problemas["resumen_facultades"][facultad]["personas_link_invalido"].append(persona_data)
                           resumen_problemas["resumen_facultades"][facultad]["carreras"][carrera]["personas"].append(persona_data)
                           resumen_problemas["links_invalidos"].append({
                                  "nombre": nombre if nombre else "SIN NOMBRE", "dni": dni,
                                  "facultad": facultad, "carrera": carrera, "motivo": motivo_invalido
                           })
                           # Link inválido/EN CONSTRUCCIÓN â†’ guardar para búsqueda automática en legajos
                           if nombre or dni:
                                  resumen_problemas.setdefault("datos_sin_link", []).append({
                                         "nombre": nombre, "dni": dni, "facultad": facultad, "carrera": carrera
                                  })
             
             pass  # log removed (encoding issue)
             print(f"   Con DNI: {sum(1 for d in datos_excel if d['dni'])}")
             print(f"   Con nombre: {sum(1 for d in datos_excel if d['nombre'])}")
             pass  # log removed (encoding issue)
             
      except Exception as e:
             print(f"[ERR] Error leyendo archivo desde memoria: {e}")
             import traceback
             traceback.print_exc()
      
      return urls, datos_excel, resumen_problemas


def leer_datos_desde_requerimientos() -> tuple:
      """
      Lee URLs, DNIs, nombres, facultades y carreras.
      
      PRIORIDAD:
      1. Archivo subido por el usuario (guardado en disco)
      2. Archivos de la carpeta LINKS como fallback
      
      Returns:
             tuple: (urls, datos_excel, resumen_problemas) donde:
                    - datos_excel es lista de {url, dni, nombre, facultad, carrera}
                    - resumen_problemas contiene info de links inválidos y datos faltantes por facultad/carrera
      """
      global archivo_excel_subido, archivo_contenido_memoria
      
      print(f"\n{'='*60}")
      pass  # log removed (encoding issue)
      print(f"{'='*60}")
      print(f"   - ARCHIVO_SUBIDO_PATH: {ARCHIVO_SUBIDO_PATH}")
      print(f"   - Existe archivo subido: {os.path.exists(ARCHIVO_SUBIDO_PATH)}")
      print(f"   - Existe carpeta LINKS: {os.path.exists(INFORMACION_VALIDACION_PATH)}")
      
      # PRIORIDAD 1: Usar archivo subido por el usuario (guardado en disco)
      if os.path.exists(ARCHIVO_SUBIDO_PATH):
             pass  # log removed (encoding issue)
             print(f"   Ruta: {ARCHIVO_SUBIDO_PATH}")
             urls, datos, resumen_problemas = leer_datos_desde_archivo_subido(ARCHIVO_SUBIDO_PATH)
             print(f"   -> URLs encontradas: {len(urls)}")
             print(f"{'='*60}\n")
             return urls, datos, resumen_problemas
      
      # PRIORIDAD 2: Usar archivos de la carpeta LINKS como fallback
      if os.path.exists(INFORMACION_VALIDACION_PATH):
             pass  # log removed (encoding issue)
             print(f"{'='*60}\n")
             return leer_datos_desde_links_usil()
      
      # Si no hay nada, retornar vacío
      print("[!] No se encontraron archivos de datos.")
      print(f"{'='*60}\n")
      return [], [], {"links_invalidos": [], "datos_faltantes": [], "resumen_facultades": {}, "resumen_carreras": {}}


def leer_urls_desde_excel() -> list:
      """Lee todas las URLs de la columna A del Excel (LEGACY - mantener compatibilidad)"""
      urls, _, _ = leer_datos_desde_requerimientos()
      return urls


def normalizar_nombre(nombre: str) -> str:
      """
      Normaliza un nombre para comparación (elimina tildes, espacios extras, etc.)
      """
      import unicodedata
      if not nombre:
             return ""
      
      # Convertir a minúsculas
      nombre = nombre.lower().strip()
      
      # Eliminar tildes y caracteres especiales
      nombre = unicodedata.normalize('NFD', nombre)
      nombre = ''.join(c for c in nombre if unicodedata.category(c) != 'Mn')
      
      # Eliminar caracteres no alfanuméricos excepto espacios
      nombre = ''.join(c if c.isalnum() or c.isspace() else ' ' for c in nombre)
      
      # Normalizar espacios múltiples
      nombre = ' '.join(nombre.split())
      
      return nombre


def combinar_cvs_pdf_web(cvs_pdf: list, cvs_web: list) -> list:
      """
      Combina datos de CVs en PDF con datos de CTI Vitae (Web).
      Hace match por nombre y toma el mejor valor de cada campo.
      
      Args:
             cvs_pdf: Lista de CVs extraídos de archivos PDF
             cvs_web: Lista de CVs extraídos de CTI Vitae
             
      Returns:
             Lista de CVs combinados con información complementada
      """
      if not cvs_pdf:
             return cvs_web
      if not cvs_web:
             return cvs_pdf
      
      print(f"\n{'='*60}")
      pass  # log removed (encoding issue)
      print(f"{'='*60}")
      
      # Crear índices de búsqueda por nombre normalizado
      indice_pdf = {}
      for cv in cvs_pdf:
             nombre_norm = normalizar_nombre(cv.get('nombre', ''))
             if nombre_norm:
                    indice_pdf[nombre_norm] = cv
      
      indice_web = {}
      for cv in cvs_web:
             nombre_norm = normalizar_nombre(cv.get('nombre', ''))
             if nombre_norm:
                    indice_web[nombre_norm] = cv
      
      cvs_combinados = []
      nombres_procesados = set()
      combinaciones_exitosas = 0
      
      # Procesar todos los CVs web (son la fuente principal)
      for cv_web in cvs_web:
             nombre_norm = normalizar_nombre(cv_web.get('nombre', ''))
             nombres_procesados.add(nombre_norm)
             
             # Buscar match en PDFs
             cv_pdf = None
             if nombre_norm in indice_pdf:
                    cv_pdf = indice_pdf[nombre_norm]
             else:
                    # Buscar coincidencia parcial (apellido)
                    for nombre_pdf, cv_p in indice_pdf.items():
                           # Verificar si comparten al menos 2 palabras (probablemente apellidos)
                           palabras_web = set(nombre_norm.split())
                           palabras_pdf = set(nombre_pdf.split())
                           coincidencias = palabras_web.intersection(palabras_pdf)
                           if len(coincidencias) >= 2:
                                  cv_pdf = cv_p
                                  break
             
             if cv_pdf:
                    # COMBINAR: Tomar el mejor valor de cada campo
                    cv_combinado = combinar_datos_cv(cv_pdf, cv_web)
                    cvs_combinados.append(cv_combinado)
                    combinaciones_exitosas += 1
                    print(f"   [OK] MATCH: {cv_web.get('nombre', 'N/A')}")
                    print(f"        PDF: {cv_pdf.get('educacion', {})} | Web: {cv_web.get('educacion', {})}")
                    print(f"        -> Combinado: {cv_combinado.get('educacion', {})}")
             else:
                    # No hay match, usar solo el CV web
                    cv_web['fuente_datos'] = 'CTI_VITAE_SOLO'
                    cvs_combinados.append(cv_web)
      
      # Agregar CVs de PDF que no tuvieron match
      for nombre_pdf, cv_pdf in indice_pdf.items():
             if nombre_pdf not in nombres_procesados:
                    cv_pdf['fuente_datos'] = 'PDF_SOLO'
                    cvs_combinados.append(cv_pdf)
                    print(f"   [PDF] Solo PDF (sin match web): {cv_pdf.get('nombre', 'N/A')}")
      
      pass  # log removed (encoding issue)
      print(f"   Total CVs resultantes: {len(cvs_combinados)}")
      print(f"   Combinaciones exitosas (PDF+Web): {combinaciones_exitosas}")
      print(f"   Solo CTI Vitae: {len([c for c in cvs_combinados if c.get('fuente_datos') == 'CTI_VITAE_SOLO'])}")
      print(f"   Solo PDF: {len([c for c in cvs_combinados if c.get('fuente_datos') == 'PDF_SOLO'])}")
      
      return cvs_combinados


def combinar_datos_cv(cv_pdf: dict, cv_web: dict) -> dict:
      """
      Combina dos CVs tomando el mejor valor de cada campo.
      Prioriza el dato más completo o el grado académico más alto.
      
      Args:
             cv_pdf: CV extraído del PDF
             cv_web: CV extraído de CTI Vitae
             
      Returns:
             CV combinado con los mejores datos de ambas fuentes
      """
      cv_combinado = {}
      
      # Campos básicos - preferir el que tenga datos
      cv_combinado['nombre'] = cv_web.get('nombre') or cv_pdf.get('nombre', '')
      cv_combinado['dni'] = cv_web.get('dni') or cv_pdf.get('dni', '')
      cv_combinado['facultad'] = cv_web.get('facultad') or cv_pdf.get('facultad', '')
      cv_combinado['url'] = cv_web.get('url', '')
      
      # Educación - TOMAR EL GRADO MÁS ALTO
      edu_pdf = cv_pdf.get('educacion', {})
      edu_web = cv_web.get('educacion', {})
      
      if isinstance(edu_pdf, dict) and isinstance(edu_web, dict):
             cv_combinado['educacion'] = {
                    'doctorado': edu_pdf.get('doctorado', False) or edu_web.get('doctorado', False),
                    'maestria': edu_pdf.get('maestria', False) or edu_web.get('maestria', False),
                    'licenciatura': edu_pdf.get('licenciatura', False) or edu_web.get('licenciatura', False)
             }
      else:
             cv_combinado['educacion'] = edu_web if edu_web else edu_pdf
      
      # Experiencia - TOMAR EL MAYOR
      cv_combinado['anos_experiencia'] = max(
             cv_pdf.get('anos_experiencia', 0) or 0,
             cv_web.get('anos_experiencia', 0) or 0
      )
      
      # Experiencia docente - TOMAR EL MAYOR
      cv_combinado['experiencia_docente'] = max(
             cv_pdf.get('experiencia_docente', 0) or 0,
             cv_web.get('experiencia_docente', 0) or 0
      )
      
      # Publicaciones - TOMAR EL MAYOR
      cv_combinado['publicaciones'] = max(
             cv_pdf.get('publicaciones', 0) or 0,
             cv_web.get('publicaciones', 0) or 0
      )
      
      # Proyectos - TOMAR EL MAYOR
      cv_combinado['proyectos'] = max(
             cv_pdf.get('proyectos', 0) or 0,
             cv_web.get('proyectos', 0) or 0
      )
      
      # Experiencia laboral - COMBINAR AMBAS
      exp_pdf = cv_pdf.get('experiencia_laboral', []) or []
      exp_web = cv_web.get('experiencia_laboral', []) or []
      cv_combinado['experiencia_laboral'] = exp_pdf + exp_web
      
      # Idiomas - COMBINAR (eliminar duplicados)
      idiomas_pdf = set(cv_pdf.get('idiomas', []) or [])
      idiomas_web = set(cv_web.get('idiomas', []) or [])
      cv_combinado['idiomas'] = list(idiomas_pdf.union(idiomas_web))
      
      # Premios - COMBINAR
      premios_pdf = cv_pdf.get('premios', []) or []
      premios_web = cv_web.get('premios', []) or []
      cv_combinado['premios'] = premios_pdf + premios_web
      
      # Texto completo - COMBINAR para análisis
      texto_pdf = cv_pdf.get('texto_completo', '') or ''
      texto_web = cv_web.get('texto_completo', '') or ''
      cv_combinado['texto_completo'] = f"{texto_pdf}\n\n=== DATOS CTI VITAE ===\n\n{texto_web}"
      
      # Detalle de publicaciones (del web si existe)
      cv_combinado['publicaciones_detalle'] = cv_web.get('publicaciones_detalle', {})
      
      # Información de estado del perfil
      cv_combinado['fecha_actualizacion'] = cv_web.get('fecha_actualizacion')
      cv_combinado['meses_sin_actualizar'] = cv_web.get('meses_sin_actualizar')
      cv_combinado['perfil_desactualizado'] = cv_web.get('perfil_desactualizado', False)
      cv_combinado['perfil_vacio'] = cv_web.get('perfil_vacio', False)
      
      # Marcar fuente como combinada
      cv_combinado['fuente'] = 'PDF+CTI_VITAE'
      cv_combinado['fuente_datos'] = 'COMBINADO'
      cv_combinado['archivo'] = cv_pdf.get('archivo', '')
      
      return cv_combinado


def verificar_conectividad_cti() -> dict:
      """
      Verifica si CTI Vitae está accesible.
      
      Returns:
             dict con: accesible (bool), mensaje (str), tiempo_respuesta (float)
      """
      import requests
      url_test = 'https://ctivitae.concytec.gob.pe/appDirectorioCTI/'
      
      try:
             inicio = time.time()
             response = requests.get(url_test, timeout=15)
             tiempo = time.time() - inicio
             
             if response.status_code == 200:
                    return {
                           "accesible": True,
                           "mensaje": f"CTI Vitae accesible (tiempo: {tiempo:.1f}s)",
                           "tiempo_respuesta": tiempo
                    }
             else:
                    return {
                           "accesible": False,
                           "mensaje": f"CTI Vitae respondió con error HTTP {response.status_code}",
                           "tiempo_respuesta": tiempo
                    }
      except requests.Timeout:
             return {
                    "accesible": False,
                    "mensaje": "CTI Vitae no responde (timeout - el sitio puede estar caído o lento)",
                    "tiempo_respuesta": None
             }
      except requests.ConnectionError:
             return {
                    "accesible": False,
                    "mensaje": "No se puede conectar a CTI Vitae (verifica tu conexión a internet)",
                    "tiempo_respuesta": None
             }
      except Exception as e:
             return {
                    "accesible": False,
                    "mensaje": f"Error verificando CTI Vitae: {str(e)}",
                    "tiempo_respuesta": None
             }


def ejecutar_evaluacion():
       """Ejecuta el proceso completo de evaluación con actualizaciones de estado y paralelismo"""
       global estado_proceso, carpeta_pdfs_subida, modo_rapido_activo, links_manuales_cti

       try:
             # Reiniciar el progreso de extracción
             reset_extraction_progress()
             
             # Registrar tiempo de inicio
             estado_proceso["tiempo_inicio"] = time.time()
             
             # PASO 1: Extracción de CVs desde PDF (SOLO si se subió carpeta de PDFs)
             cvs_procesados = []
             cvs_pdf = []
             
             if carpeta_pdfs_subida:
                    estado_proceso.update({
                           "paso_actual": 1,
                           "mensaje": " Preparando análisis de perfiles (PDF)...",
                           "porcentaje": 10,
                           "cvs_procesados": 0,
                           "cvs_total": 0
                    })
                    time.sleep(0.5)
                    
                    def cb_progreso_pdf(archivo, completados, total):
                        porcentaje_actual = 10 + int((completados / max(total, 1)) * 30) # del 10% al 40%
                        estado_proceso.update({
                            "mensaje": f" Analizando PDF [{completados}/{total}]: {archivo}",
                            "porcentaje": porcentaje_actual,
                            "cvs_procesados": completados,
                            "cvs_total": total
                        })
                    
                    cvs_procesados = procesar_todos_cvs(carpeta_pdfs_subida, callback_progreso=cb_progreso_pdf)
                    print(f"[DIR] Usando carpeta de PDFs subida: {carpeta_pdfs_subida}")
                    pass  # log removed (encoding issue)
                    
                    # Guardar CVs de PDF para combinar después
                    cvs_pdf = cvs_procesados.copy()
             else:
                    pass  # log removed (encoding issue)
                    estado_proceso.update({
                           "paso_actual": 1,
                           "mensaje": " Preparando extracción de perfiles CTI Vitae...",
                           "porcentaje": 5,
                           "cvs_procesados": 0,
                           "cvs_total": 0
                    })
                    time.sleep(0.3)
             
             # PASO 1.5: Extracción de CVs desde web (leer URLs + DNI + nombres + facultades del Excel)
             urls_web, datos_excel, resumen_problemas = [], [], {
                    "links_invalidos": [],
                    "datos_faltantes": [],
                    "resumen_facultades": {},
                    "datos_sin_link": []
             }
             
             usar_excel = estado_proceso.get("usar_excel", True)
             
             if not usar_excel:
                    print("[FAST] Modo individual: omitiendo carga de Excel por defecto")
                    estado_proceso.update({
                           "paso_actual": 1,
                           "mensaje": ">> Modo individual: evaluando solo PDFs cargados",
                           "porcentaje": 20
                    })
             elif modo_rapido_activo:
                    print("[FAST] Modo rapido activo: se omite lectura/extraccion CTI")
                    estado_proceso.update({
                           "paso_actual": 1,
                           "mensaje": ">> Modo rapido: solo fuentes cargadas manualmente",
                           "porcentaje": 20
                    })
             else:
                    urls_web, datos_excel, resumen_problemas = leer_datos_desde_requerimientos()

             # PASO 1.5b: Sumar los links de CTI Vitae agregados manualmente desde el panel
             if links_manuales_cti:
                    urls_existentes = set(urls_web)
                    for url_manual in links_manuales_cti:
                           if url_manual in urls_existentes:
                                  continue
                           urls_web.append(url_manual)
                           datos_excel.append({
                                  'url': url_manual, 'dni': '', 'nombre': '',
                                  'facultad': '', 'carrera': ''
                           })
                           urls_existentes.add(url_manual)
                    print(f"[LINKS MANUALES] +{len(links_manuales_cti)} link(s) agregados desde el panel")

             cvs_web = []
             cti_no_disponible = False
             if urls_web:
                    total_urls = len(urls_web)
                    
                    # VERIFICAR CONECTIVIDAD CON CTI VITAE ANTES DE INICIAR
                    estado_proceso.update({
                           "paso_actual": 1,
                           "mensaje": " Verificando acceso a CTI Vitae...",
                           "porcentaje": 8
                    })
                    time.sleep(0.3)
                    
                    verificacion_cti = verificar_conectividad_cti()
                    if not verificacion_cti["accesible"]:
                           print(f"[!] CTI VITAE NO DISPONIBLE: {verificacion_cti['mensaje']}")
                           cti_no_disponible = True
                           estado_proceso.update({
                                  "paso_actual": 1,
                                  "mensaje": f" {verificacion_cti['mensaje']} - Continuando con datos del Excel...",
                                  "porcentaje": 10
                           })
                           time.sleep(1)
                           
                           # Crear CVs vacíos con datos del Excel para mostrar al menos los nombres
                           for i, dato in enumerate(datos_excel):
                                  cv_vacio = {
                                         'nombre': dato.get('nombre', 'Sin nombre'),
                                         'dni': dato.get('dni', ''),
                                         'facultad': dato.get('facultad', ''),
                                         'carrera': dato.get('carrera', ''),
                                         'url': dato.get('url', ''),
                                         'anos_experiencia': 0,
                                         'educacion': {'doctorado': False, 'maestria': False, 'licenciatura': False},
                                         'publicaciones': 0,
                                         'proyectos': 0,
                                         'experiencia_laboral': [],
                                         'experiencia_docente': 0,
                                         'idiomas': [],
                                         'premios': [],
                                         'error_extraccion': True,
                                         'error_mensaje': verificacion_cti['mensaje'],
                                         'perfil_vacio': True,
                                         'cti_no_disponible': True,
                                         'indice_original': i + 1
                                  }
                                  cvs_web.append(cv_vacio)
                                  estado_proceso["registros_con_error"].append({
                                         "indice": i + 1,
                                         "url": dato.get('url', ''),
                                         "nombre": dato.get('nombre', 'Sin nombre'),
                                         "dni": dato.get('dni', ''),
                                         "facultad": dato.get('facultad', ''),
                                         "carrera": dato.get('carrera', ''),
                                         "error": f"CTI Vitae no disponible: {verificacion_cti['mensaje']}"
                                  })
                    else:
                           print(f"[OK] CTI Vitae accesible - {verificacion_cti['mensaje']}")
                    
                    # Solo hacer extracción si CTI está disponible
                    if not cti_no_disponible:
                           pass  # log removed (encoding issue)
                           estado_proceso.update({
                                  "paso_actual": 1,
                                  "mensaje": f">> Extrayendo {total_urls} perfiles CTI (paralelo)...",
                                  "porcentaje": 15,
                                  "cvs_total": total_urls,
                                  "cvs_procesados": 0
                           })
                           time.sleep(0.5)
                           
                           # Función callback para actualizar progreso en tiempo real
                           def actualizar_progreso_extraccion(progreso):
                                  completados = progreso.get("completados", 0)
                                  total = progreso.get("total", 1)
                                  porcentaje_extraccion = int((completados / total) * 25) + 15  # 15-40%
                                  
                                  # Calcular tiempo estimado
                                  tiempo_restante_str = ""
                                  if progreso.get("tiempo_inicio") and completados > 0:
                                         tiempo_transcurrido = time.time() - progreso["tiempo_inicio"]
                                         velocidad = completados / tiempo_transcurrido
                                         if velocidad > 0 and total > completados:
                                                restante = (total - completados) / velocidad
                                                if restante < 60:
                                                       tiempo_restante_str = f" (~{int(restante)}s restantes)"
                                                else:
                                                       mins = int(restante // 60)
                                                       segs = int(restante % 60)
                                                       tiempo_restante_str = f" (~{mins}m {segs}s restantes)"
                                  
                                  estado_proceso.update({
                                         "mensaje": f">> Extrayendo CVs: {completados}/{total}{tiempo_restante_str}",
                                         "porcentaje": min(porcentaje_extraccion, 40),
                                         "cvs_procesados": completados,
                                         "cvs_exitosos": progreso.get("exitosos", 0),
                                         "cvs_con_error": progreso.get("errores", 0)
                                  })
                           
                           extractor_web = ExtractorWebCVs()
                           # Usar extracción paralela con 10 workers para mayor velocidad
                           cvs_web = extractor_web.extraer_multiples_cvs(
                                  urls_web, 
                                  datos_excel, 
                                  max_workers=10,
                                  callback_progreso=actualizar_progreso_extraccion
                           )
                           
                           # Guardar registros con error para permitir re-análisis
                           progreso_final = get_extraction_progress()
                           estado_proceso["registros_con_error"] = progreso_final.get("registros_con_error", [])
                    
                    pass  # log removed (encoding issue)
             else:
                    print("[!] No se encontraron URLs en el Excel para procesar")
             
             # NUEVO: Combinar datos de PDF y Web si hay ambos
             if cvs_pdf and cvs_web:
                    estado_proceso.update({
                           "paso_actual": 1,
                           "mensaje": f" Combinando {len(cvs_pdf)} PDFs con {len(cvs_web)} perfiles web...",
                           "porcentaje": 42
                    })
                    time.sleep(0.5)
                    cvs_procesados = combinar_cvs_pdf_web(cvs_pdf, cvs_web)
             elif cvs_web:
                    cvs_procesados = cvs_web
             # Si solo hay PDF, ya están en cvs_procesados
             
             # â" PASO 1.6: Legajos (manual por botones, no automático) â"
             cvs_legajos = []
             datos_sin_link = resumen_problemas.get("datos_sin_link", [])
             if datos_sin_link:
                    if AUTO_INDEXACION_LEGAJOS:
                           estado_proceso.update({
                                  "paso_actual": 1,
                                  "mensaje": f" Indexando carpetas de legajos...",
                                  "porcentaje": 44
                           })
                           print(f"\n{'='*60}")
                           pass  # log removed (encoding issue)
                           print(f"{'='*60}")
                           indice_legajos = _construir_indice_legajos()
                           estado_proceso.update({
                                  "paso_actual": 1,
                                  "mensaje": f" Buscando {len(datos_sin_link)} CVs físicos en legajos...",
                                  "porcentaje": 45
                           })
                           total_sin_link = len(datos_sin_link)
                           for i, candidato in enumerate(datos_sin_link):
                                  pct_legajo = 45 + int((i / max(total_sin_link, 1)) * 15)
                                  estado_proceso.update({
                                         "paso_actual": 1,
                                         "mensaje": f" Legajos [{i+1}/{total_sin_link}]: buscando {candidato.get('nombre','')[:30]}...",
                                         "porcentaje": pct_legajo
                                  })
                                  nombre_c = candidato.get("nombre", "")
                                  dni_c = candidato.get("dni", "")
                                  if not nombre_c and not dni_c:
                                         continue
                                  print(f"   [{i+1}/{len(datos_sin_link)}] [?] {nombre_c} (DNI: {dni_c})")
                                  try:
                                         resultado = buscar_cv_en_legajos(nombre_c, dni_c, indice=indice_legajos)
                                         if resultado.get("encontrado") and resultado.get("cv_data"):
                                                cv = resultado["cv_data"]
                                                if not cv.get("nombre") or cv["nombre"] in ["Sin nombre", "Nombre no encontrado", ""]:
                                                       cv["nombre"] = nombre_c
                                                cv["dni"] = cv.get("dni") or dni_c
                                                cv["facultad"] = candidato.get("facultad", "")
                                                cv["carrera"] = candidato.get("carrera", "")
                                                cv["fuente_datos"] = "LEGAJO_PDF"
                                                cv["url"] = ""
                                                cvs_legajos.append(cv)
                                                print(f"       [OK] CV encontrado: {resultado.get('nombre_carpeta', '')}")
                                         else:
                                                cvs_legajos.append({
                                                       "nombre": nombre_c if nombre_c else "SIN NOMBRE",
                                                       "dni": dni_c,
                                                       "facultad": candidato.get("facultad", ""),
                                                       "carrera": candidato.get("carrera", ""),
                                                       "anos_experiencia": 0,
                                                       "experiencia_docente": 0,
                                                       "educacion": {"doctorado": False, "maestria": False, "licenciatura": False, "bachiller": False},
                                                       "publicaciones": 0,
                                                       "proyectos_investigacion": 0,
                                                       "idiomas": [],
                                                       "fuente_datos": "SIN_CV",
                                                       "error_extraccion": True,
                                                       "error_mensaje": resultado.get("mensaje", "CV no encontrado en legajos"),
                                                       "sin_link_cti": True,
                                                       "url": ""
                                                })
                                                print(f"       [ERR] No encontrado: {resultado.get('mensaje', '')}")
                                  except Exception as e_leg:
                                         print(f"       [!] Error buscando en legajos: {e_leg}")
                                         cvs_legajos.append({
                                                "nombre": nombre_c if nombre_c else "SIN NOMBRE",
                                                "dni": dni_c,
                                                "facultad": candidato.get("facultad", ""),
                                                "carrera": candidato.get("carrera", ""),
                                                "anos_experiencia": 0, "experiencia_docente": 0,
                                                "educacion": {"doctorado": False, "maestria": False, "licenciatura": False, "bachiller": False},
                                                "publicaciones": 0, "fuente_datos": "SIN_CV",
                                                "error_extraccion": True, "error_mensaje": str(e_leg), "url": ""
                                         })

                           cvs_procesados = cvs_procesados + cvs_legajos
                           encontrados = sum(1 for c in cvs_legajos if not c.get("error_extraccion"))
                           print(f"\n[OK] PASO 1.6 completado: {encontrados}/{len(cvs_legajos)} CVs encontrados en legajos")
                           estado_proceso.update({
                                  "paso_actual": 1,
                                  "mensaje": f"Legajos: {encontrados}/{len(cvs_legajos)} CVs encontrados",
                                  "porcentaje": 60
                           })
                    else:
                           print(f"[DIR] PASO 1.6 desactivado (manual): {len(datos_sin_link)} candidatos sin link CTI pendientes")
                           estado_proceso.update({
                                  "paso_actual": 1,
                                  "mensaje": f" {len(datos_sin_link)} candidatos sin link CTI (usar botones Legajos/Web)",
                                  "porcentaje": 60
                           })

                           # Mantener visibles los candidatos pendientes para activarlos manualmente por botones.
                           placeholders = []
                           for candidato in datos_sin_link:
                                  placeholders.append({
                                         "nombre": candidato.get("nombre", "SIN NOMBRE") or "SIN NOMBRE",
                                         "dni": candidato.get("dni", ""),
                                         "facultad": candidato.get("facultad", ""),
                                         "carrera": candidato.get("carrera", ""),
                                         "url": "",
                                         "anos_experiencia": 0,
                                         "experiencia_docente": 0,
                                         "educacion": {"doctorado": False, "maestria": False, "licenciatura": False, "bachiller": False},
                                         "publicaciones": 0,
                                         "proyectos_investigacion": 0,
                                         "idiomas": [],
                                         "fuente_datos": "SIN_CV",
                                         "error_extraccion": True,
                                         "error_mensaje": "Pendiente de reconstrucción manual (Legajos/Web)",
                                         "sin_link_cti": True,
                                         "perfil_vacio": True,
                                  })
                           cvs_legajos = placeholders
                           cvs_procesados = cvs_procesados + placeholders
             # -- FIN PASO 1.6 --

             # â" PASO 1.7: Búsqueda web experimental (desactivada por defecto) â"
             if AUTO_RECONSTRUCCION_WEB_EXPERIMENTAL:
                    candidatos_sin_cv = [c for c in cvs_legajos if c.get("fuente_datos") == "SIN_CV"]
                    if candidatos_sin_cv:
                           try:
                                  from buscador_web_cv import buscar_cv_en_web as _buscar_web
                                  estado_proceso.update({
                                         "paso_actual": 1,
                                         "mensaje": f" Buscando perfiles web (experimental)...",
                                         "porcentaje": 62
                                  })
                                  print(f"\n{'='*60}")
                                  pass  # log removed (encoding issue)
                                  print(f"{'='*60}")
                                  for i, c_sin_cv in enumerate(candidatos_sin_cv):
                                         nombre_w  = c_sin_cv.get("nombre", "")
                                         dni_w     = c_sin_cv.get("dni", "")
                                         fac_w     = c_sin_cv.get("facultad", "")
                                         car_w     = c_sin_cv.get("carrera", "")
                                         if not nombre_w:
                                                continue
                                         pct = 62 + int((i / max(len(candidatos_sin_cv), 1)) * 8)
                                         estado_proceso.update({
                                                "mensaje": f" Web [{i+1}/{len(candidatos_sin_cv)}]: {nombre_w[:30]}...",
                                                "porcentaje": pct
                                         })
                                         try:
                                                res_web = _buscar_web(nombre_w, dni_w, fac_w, car_w)
                                                if res_web.get("encontrado") and res_web.get("cv_data"):
                                                       cv_web = res_web["cv_data"]
                                                       cv_web["nombre"] = nombre_w
                                                       cv_web["dni"] = dni_w or cv_web.get("dni", "")
                                                       cv_web["facultad"] = fac_w
                                                       cv_web["carrera"] = car_w
                                                       for j, cp in enumerate(cvs_procesados):
                                                              if _normalizar_para_match(cp.get("nombre", "")) == _normalizar_para_match(nombre_w):
                                                                     cvs_procesados[j] = cv_web
                                                                     break
                                                       print(f"       [OK] Web: {nombre_w} | confianza={cv_web.get('score_confianza', 0)}%")
                                                else:
                                                       print(f"       [--] Sin datos web: {nombre_w}")
                                         except Exception as e_web:
                                                print(f"       [!] Error web {nombre_w}: {e_web}")
                           except ImportError:
                                  pass  # log removed (encoding issue)
             else:
                    pass  # log removed (encoding issue)
             # -- FIN PASO 1.7 --

             if not cvs_procesados:
                    # Intentar cargar datos de prueba si no hay CVs
                    datos_prueba_path = os.path.join(os.path.dirname(BASE_DIR), "Testear", "datos_para_ia.json")
                    if os.path.exists(datos_prueba_path):
                           estado_proceso.update({
                                  "paso_actual": 1,
                                  "mensaje": " Cargando datos de prueba (demo)...",
                                  "porcentaje": 30
                           })
                           time.sleep(0.5)
                           
                           import json
                           with open(datos_prueba_path, 'r', encoding='utf-8') as f:
                                  datos_demo = json.load(f)
                           
                           # Convertir datos de prueba al formato esperado
                           evaluaciones = []
                           for candidato in datos_demo.get('candidatos', []):
                                  eval_data = candidato.get('evaluacion', {})
                                  evaluacion = {
                                         'nombre': candidato.get('nombre_completo', 'Sin nombre'),
                                         'total': eval_data.get('puntaje_total', 0),
                                         'maximo': eval_data.get('puntaje_maximo_posible', 200),
                                         'porcentaje': eval_data.get('porcentaje_general', 0),
                                         'clasificacion': eval_data.get('clasificacion', 'SIN_CLASIFICAR'),
                                         'es_elegible': eval_data.get('es_elegible', False),
                                         'tipo_perfil': eval_data.get('tipo_perfil', 'general'),
                                         'puntajes': {
                                                'C1': eval_data.get('C1_formacion_academica', {}).get('puntaje_asignado', 0),
                                                'C2': eval_data.get('C2_experiencia_docente', {}).get('puntaje_asignado', 0),
                                                'C3': eval_data.get('C3_experiencia_profesional', {}).get('puntaje_asignado', 0),
                                                'C4': eval_data.get('C4_centro_labores', {}).get('puntaje_asignado', 0),
                                                'C5': eval_data.get('C5_produccion_academica', {}).get('puntaje_asignado', 0),
                                                'C6': eval_data.get('C6_liderazgo', {}).get('puntaje_asignado', 0),
                                                'C7': eval_data.get('C7_especializacion', {}).get('puntaje_asignado', 0)
                                         },
                                         'detalles': {
                                                'C1': eval_data.get('C1_formacion_academica', {}).get('detalle', {}),
                                                'C2': eval_data.get('C2_experiencia_docente', {}).get('detalle', {}),
                                                'C3': eval_data.get('C3_experiencia_profesional', {}).get('detalle', {}),
                                                'C4': eval_data.get('C4_centro_labores', {}).get('detalle', {}),
                                                'C5': eval_data.get('C5_produccion_academica', {}).get('detalle', {})
                                         },
                                         'archivo': 'datos_demo',
                                         'fuente': candidato.get('fuente_datos', 'Demo')
                                  }
                                  evaluaciones.append(evaluacion)
                           
                           # Saltar directamente a generar resultados
                           estado_proceso.update({
                                  "paso_actual": 3,
                                  "mensaje": " Generando insights y ranking de talento...",
                                  "porcentaje": 80
                           })
                           time.sleep(1)
                           
                           # Guardar JSON
                           timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                           archivo_json = os.path.join(RESULTADOS_DIR, f"clasificacion_final_{timestamp}.json")
                           
                           output = {
                                  "metadata": {
                                         "version": "3.0",
                                         "fecha": datetime.now().isoformat(),
                                         "total_candidatos": len(evaluaciones),
                                         "modo": "DEMO"
                                  },
                                  "ranking": evaluaciones
                           }
                           
                           with open(archivo_json, "w", encoding="utf-8") as f:
                                  json.dump(output, f, indent=2, ensure_ascii=False)
                           
                           resultado = {
                                  "evaluaciones": evaluaciones,
                                  "reportes": {"json": os.path.basename(archivo_json)},
                                  "fecha": datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                                  "modo": "DEMO"
                           }
                           
                           estado_proceso.update({
                                  "paso_actual": 3,
                                  "mensaje": " Evaluación completada (MODO DEMO)",
                                  "porcentaje": 100,
                                  "completado": True,
                                  "resultado": resultado
                           })
                           return
                    else:
                           estado_proceso.update({
                                  "error": "No se encontraron CVs para procesar. Coloque archivos PDF en la carpeta 'Cvs' o configure URLs en el Excel de LINKS.",
                                  "completado": True
                           })
                           return
             
             # PASO 2: Evaluación con nueva lógica simplificada
             estado_proceso.update({
                    "paso_actual": 2,
                    "mensaje": f" Evaluando {len(cvs_procesados)} perfiles con IA...",
                    "porcentaje": 50
             })
             time.sleep(1)
             
             motor = MotorEvaluacion()
             evaluaciones = motor.evaluar_multiples_cvs(cvs_procesados)
             
             # PASO 3: Preparar resultados
             estado_proceso.update({
                    "paso_actual": 3,
                    "mensaje": " Generando insights y ranking de talento...",
                    "porcentaje": 80
             })
             time.sleep(1)
             
             # Guardar JSON
             import json
             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
             archivo_json = os.path.join(RESULTADOS_DIR, f"clasificacion_final_{timestamp}.json")
             
             output = {
                    "metadata": {
                           "version": "3.0",
                           "fecha": datetime.now().isoformat(),
                           "total_candidatos": len(evaluaciones)
                    },
                    "ranking": evaluaciones
             }

             with open(archivo_json, "w", encoding="utf-8") as f:
                    json.dump(output, f, indent=2, ensure_ascii=False)
             
             # Preparar resultado para mostrar en web
             # NUEVO: Detectar perfiles con problemas (desactualizados, vacíos)
             perfiles_con_problemas = []
             for e in evaluaciones:
                    problemas_perfil = []
                    
                    if e.get('perfil_desactualizado'):
                           problemas_perfil.append('DESACTUALIZADO')
                    if e.get('perfil_vacio'):
                           problemas_perfil.append('PERFIL VACÍO')
                    if e.get('meses_sin_actualizar') and e.get('meses_sin_actualizar') > 6:
                           problemas_perfil.append(f'+{e.get("meses_sin_actualizar")} meses sin actualizar')
                    
                    if problemas_perfil:
                           perfiles_con_problemas.append({
                                  "nombre": e.get('nombre', ''),
                                  "dni": e.get('dni', ''),
                                  "facultad": e.get('facultad', ''),
                                  "motivo": ' | '.join(problemas_perfil),
                                  "meses_sin_actualizar": e.get('meses_sin_actualizar'),
                                  "fecha_actualizacion": e.get('fecha_actualizacion')
                           })
                           
                           # Agregar también a resumen_problemas para mostrar en la tabla
                           resumen_problemas.setdefault("perfiles_desactualizados", []).append({
                                  "nombre": e.get('nombre', ''),
                                  "dni": e.get('dni', ''),
                                  "facultad": e.get('facultad', ''),
                                  "motivo": problemas_perfil[0] if problemas_perfil else 'PROBLEMA',
                                  "meses_sin_actualizar": e.get('meses_sin_actualizar')
                           })
             
             pass  # log removed (encoding issue)
             
             resultado = {
                    "evaluaciones": [
                           {
                                  "nombre": e['nombre'],
                                  "dni": e.get('dni', ''),
                                  "total": e['total'],
                                  "maximo": e['maximo'],
                                  "porcentaje": e['porcentaje'],
                                  "clasificacion": e['clasificacion'],
                                  "es_elegible": e['es_elegible'],
                                  "tipo_perfil": e.get('tipo_perfil', 'general'),
                                  "puntajes": e['puntajes'],
                                  "detalles": e['detalles'],
                                  "archivo": e.get('archivo', ''),
                                  "fuente": e.get('fuente', ''),
                                  "fuente_datos": e.get('fuente_datos', ''),
                                  "facultad": e.get('facultad', ''),
                                  "carrera": e.get('carrera', ''),
                                  # Datos de la solicitud (hoja 2026.1) — necesarios para los filtros
                                  "tipo_candidato": e.get('tipo_candidato', ''),
                                  "puesto": e.get('puesto', ''),
                                  "categoria_practitioner": e.get('categoria_practitioner', ''),
                                  # Clasificación de talento (para ranking y Excel detallado)
                                  "nivel": e.get('nivel', ''),
                                  "dominio": e.get('dominio', ''),
                                  "perfil_dominante": e.get('perfil_dominante', ''),
                                  "perfil_secundario": e.get('perfil_secundario', ''),
                                  "justificacion_perfil": e.get('justificacion_perfil', ''),
                                  "facultad_recomendada": e.get('facultad_recomendada', ''),
                                  "carrera_recomendada": e.get('carrera_recomendada', ''),
                                  "cursos_recomendados": e.get('cursos_recomendados', ''),
                                  "etiqueta_reclutamiento": e.get('etiqueta_reclutamiento', ''),
                                  # URL para re-análisis — la mayoría de fuentes la dejan en 'archivo'
                                  "url": _url_persona(e),
                                  # NUEVO: Bandera de error en extracción
                                  "error_extraccion": e.get('error_extraccion', False),
                                  "error_mensaje": e.get('error_mensaje', ''),
                                  # NUEVO: Información de estado del perfil
                                  "perfil_desactualizado": e.get('perfil_desactualizado', False),
                                  "perfil_vacio": e.get('perfil_vacio', False),
                                  "meses_sin_actualizar": e.get('meses_sin_actualizar'),
                                  "fecha_actualizacion": e.get('fecha_actualizacion'),
                                  "info_incompleta": e.get('perfil_desactualizado', False) or e.get('perfil_vacio', False) or e.get('error_extraccion', False),
                                  "justificacion_decision": e.get('justificacion_decision', '')
                           }
                           for e in evaluaciones
                    ],
                    "reportes": {
                           "json": os.path.basename(archivo_json)
                    },
                    "fecha": datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    # Agregar resumen de problemas detectados
                    "resumen_problemas": resumen_problemas,
                    "perfiles_con_problemas": perfiles_con_problemas
             }
             
             # Debug: Imprimir información de resumen_problemas
             print(f"\n[?] DEBUG resumen_problemas:")
             pass  # log removed (encoding issue)
             print(f"   Datos faltantes: {len(resumen_problemas.get('datos_faltantes', []))}")
             print(f"   Facultades: {list(resumen_problemas.get('resumen_facultades', {}).keys())}")
             
             # Imprimir resumen final en consola
             if resumen_problemas and (resumen_problemas.get("links_invalidos") or resumen_problemas.get("datos_faltantes")):
                    print("\n" + "="*80)
                    pass  # log removed (encoding issue)
                    print("="*80)
                    print(f"\n[OK] Perfiles evaluados exitosamente: {len(evaluaciones)}")
                    pass  # log removed (encoding issue)
                    print(f"[!]  Registros con datos incompletos: {len(resumen_problemas.get('datos_faltantes', []))}")
                    
                    pass  # log removed (encoding issue)
                    print("-"*80)
                    for facultad, stats in resumen_problemas.get("resumen_facultades", {}).items():
                           problemas = stats.get("link_invalido", 0) + stats.get("sin_link", 0)
                           if problemas > 0 or stats.get("datos_incompletos", 0) > 0:
                                  pass  # log removed (encoding issue)
                                  print(f"      Total registros: {stats.get('total', 0)}")
                                  pass  # log removed (encoding issue)
                                  print(f"      [ERR] Sin link: {stats.get('sin_link', 0)}")
                                  pass  # log removed (encoding issue)
                                  
                                  # Listar personas sin link
                                  if stats.get("personas_sin_link"):
                                         pass  # log removed (encoding issue)
                                         for p in stats["personas_sin_link"][:5]:
                                                print(f"           - {p.get('nombre', 'N/A')} (DNI: {p.get('dni', 'N/A')})")
                                         if len(stats["personas_sin_link"]) > 5:
                                                pass  # log removed (encoding issue)
                                  
                                  # Listar personas con link inválido
                                  if stats.get("personas_link_invalido"):
                                         pass  # log removed (encoding issue)
                                         for p in stats["personas_link_invalido"][:5]:
                                                print(f"           - {p.get('nombre', 'N/A')} (DNI: {p.get('dni', 'N/A')}) - {p.get('motivo', '')}")
                                         if len(stats["personas_link_invalido"]) > 5:
                                                pass  # log removed (encoding issue)
                    
                    print("\n" + "="*80)
             
             # Proceso completado
             estado_proceso.update({
                    "paso_actual": 3,
                    "mensaje": " Evaluación completada exitosamente",
                    "porcentaje": 100,
                    "completado": True,
                    "resultado": resultado
             })

       except Exception as e:
             estado_proceso.update({
                    "error": f"Error durante el proceso: {str(e)}",
                    "completado": True,
                    "porcentaje": 0
             })


@app.route('/')
def index():
      """Página principal"""
      return render_template('index.html')


@app.route('/subir-archivo', methods=['POST'])
def subir_archivo():
      """Recibe y valida un archivo Excel con links de CTI Vitae"""
      global archivo_excel_subido, archivo_contenido_memoria
      
      if 'archivo' not in request.files:
             return jsonify({'success': False, 'error': 'No se recibió ningún archivo'})
      
      archivo = request.files['archivo']
      
      if archivo.filename == '':
             return jsonify({'success': False, 'error': 'No se seleccionó ningún archivo'})
      
      # Validar extensión
      extension = archivo.filename.rsplit('.', 1)[-1].lower()
      if extension not in ['xlsx', 'xls']:
             return jsonify({'success': False, 'error': 'Formato no válido. Use .xlsx o .xls'})
      
      try:
             from io import BytesIO
             
             print(f"\n{'='*60}")
             pass  # log removed (encoding issue)
             print(f"{'='*60}")
             
             # Leer el contenido del archivo
             contenido_bytes = archivo.read()
             pass  # log removed (encoding issue)
             
             # IMPORTANTE: Guardar en disco para persistir entre requests
             print(f"   Guardando en: {ARCHIVO_SUBIDO_PATH}")
             with open(ARCHIVO_SUBIDO_PATH, 'wb') as f:
                    f.write(contenido_bytes)
             print(f"   [OK] Archivo guardado correctamente")
             
             archivo_excel_subido = archivo.filename
             archivo_contenido_memoria = BytesIO(contenido_bytes)
             
             print(f"\n[DIR] Archivo guardado en: {ARCHIVO_SUBIDO_PATH}")
             print(f"   Nombre original: {archivo.filename}")
             
             # Validar contenido del Excel
             urls, datos, resumen = leer_datos_desde_memoria(archivo_contenido_memoria)
             
             datos_sin_link = resumen.get('datos_sin_link', [])
             
             if len(urls) == 0 and len(resumen.get('links_invalidos', [])) == 0 and len(datos_sin_link) == 0:
                    # Eliminar archivo si no es válido
                    if os.path.exists(ARCHIVO_SUBIDO_PATH):
                           os.remove(ARCHIVO_SUBIDO_PATH)
                    archivo_contenido_memoria = None
                    archivo_excel_subido = None
                    return jsonify({
                           'success': False, 
                           'error': 'No se encontraron datos válidos en el archivo. Debe tener columnas: Nombre, DNI y Enlace CTI Vitae (o similar).'
                    })
             
             total_registros = len(urls) + len(resumen.get('links_invalidos', []))
             links_faltantes = len(resumen.get('links_invalidos', []))
             buscar_legajos = len(datos_sin_link)
             
             print(f"\n[OK] Archivo procesado correctamente: {archivo.filename}")
             print(f"   URLs encontradas: {len(urls)}")
             pass  # log removed (encoding issue)
             print(f"   Candidatos para buscar en legajos: {buscar_legajos}")
             
             # Construir mensaje descriptivo
             partes_msg = []
             if len(urls) > 0:
                    partes_msg.append(f"{len(urls)} con link CTI Vitae")
             if buscar_legajos > 0:
                    partes_msg.append(f"{buscar_legajos} se buscarán en legajos (sin link)")
             
             mensaje = f"Archivo válido: {total_registros} candidatos — " + ", ".join(partes_msg) if partes_msg else f"Archivo válido: {total_registros} registros"
             
             # Contar con DNI/nombre incluyendo datos_sin_link
             con_dni = sum(1 for d in datos if d['dni']) + sum(1 for d in datos_sin_link if d.get('dni'))
             con_nombre = sum(1 for d in datos if d['nombre']) + sum(1 for d in datos_sin_link if d.get('nombre'))
             
             return jsonify({
                    'success': True,
                    'mensaje': mensaje,
                    'urls_encontradas': len(urls),
                    'links_faltantes': links_faltantes,
                    'buscar_en_legajos': buscar_legajos,
                    'total_registros': total_registros,
                    'con_dni': con_dni,
                    'con_nombre': con_nombre
             })
             
      except Exception as e:
             print(f"[ERR] Error al procesar archivo: {e}")
             import traceback
             traceback.print_exc()
             return jsonify({'success': False, 'error': f'Error al procesar archivo: {str(e)}'})


@app.route('/quitar-archivo', methods=['POST'])
def quitar_archivo():
      """Elimina el archivo subido y vuelve al modo predeterminado"""
      global archivo_excel_subido, archivo_contenido_memoria
      
      # Eliminar archivo guardado en disco
      if os.path.exists(ARCHIVO_SUBIDO_PATH):
             try:
                    os.remove(ARCHIVO_SUBIDO_PATH)
                    pass  # log removed (encoding issue)
             except Exception as e:
                    print(f"[!] No se pudo eliminar archivo: {e}")
      
      archivo_excel_subido = None
      archivo_contenido_memoria = None
      return jsonify({'success': True, 'mensaje': 'Archivo eliminado. Se usará el archivo predeterminado.'})


@app.route('/subir-carpeta-pdfs', methods=['POST'])
def subir_carpeta_pdfs():
      """Recibe los archivos PDF de la carpeta seleccionada"""
      global carpeta_pdfs_subida
      
      if 'archivos_pdf' not in request.files:
             return jsonify({'success': False, 'error': 'No se recibieron archivos PDF'})
      
      archivos = request.files.getlist('archivos_pdf')
      nombre_carpeta = request.form.get('nombre_carpeta', 'CVs_subidos')
      
      if len(archivos) == 0:
             return jsonify({'success': False, 'error': 'No se seleccionaron archivos PDF'})
      
      try:
             # Crear carpeta temporal para los PDFs
             carpeta_destino = os.path.join(UPLOAD_FOLDER, f"pdfs_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
             os.makedirs(carpeta_destino, exist_ok=True)
             
             archivos_guardados = 0
             for archivo in archivos:
                    if archivo.filename.lower().endswith('.pdf'):
                           # Guardar con nombre seguro
                           nombre_seguro = os.path.basename(archivo.filename)
                           ruta_archivo = os.path.join(carpeta_destino, nombre_seguro)
                           archivo.save(ruta_archivo)
                           archivos_guardados += 1
             
             if archivos_guardados == 0:
                    shutil.rmtree(carpeta_destino)
                    return jsonify({
                           'success': False, 
                           'error': 'No se encontraron archivos PDF válidos'
                    })
             
             # Guardar referencia a la carpeta
             carpeta_pdfs_subida = carpeta_destino
             
             print(f"\n[OK] Carpeta de PDFs subida correctamente: {nombre_carpeta}")
             print(f"   Archivos PDF guardados: {archivos_guardados}")
             pass  # log removed (encoding issue)
             
             return jsonify({
                    'success': True,
                    'mensaje': f'{archivos_guardados} archivos PDF cargados correctamente',
                    'total_archivos': archivos_guardados,
                    'carpeta': nombre_carpeta
             })
             
      except Exception as e:
             print(f"[ERR] Error al procesar carpeta de PDFs: {e}")
             return jsonify({'success': False, 'error': f'Error al procesar archivos: {str(e)}'})


@app.route('/quitar-carpeta-pdfs', methods=['POST'])
def quitar_carpeta_pdfs():
      """Elimina la carpeta de PDFs subida"""
      global carpeta_pdfs_subida
      
      if carpeta_pdfs_subida and os.path.exists(carpeta_pdfs_subida):
             try:
                    shutil.rmtree(carpeta_pdfs_subida)
                    pass  # log removed (encoding issue)
             except Exception as e:
                    print(f"[!] No se pudo eliminar carpeta: {e}")
      
      carpeta_pdfs_subida = None
      return jsonify({'success': True, 'mensaje': 'Carpeta de PDFs eliminada. Se usará la carpeta predeterminada.'})


@app.route('/iniciar-evaluacion', methods=['POST', 'GET'])
def iniciar_evaluacion():
      """Inicia el proceso de evaluación en un hilo separado"""
      global estado_proceso, modo_rapido_activo, links_manuales_cti

      payload = request.get_json(silent=True) or {}
      modo_rapido_activo = bool(payload.get('modo_rapido', False))
      usar_excel = bool(payload.get('usar_excel', True))
      links_manuales_cti = [
             u.strip() for u in payload.get('links_manuales', [])
             if isinstance(u, str) and u.strip()
      ]
      
      print(f"\n{'='*60}")
      pass  # log removed (encoding issue)
      pass  # log removed (encoding issue)
      pass  # log removed (encoding issue)
      pass  # log removed (encoding issue)
      print(f"{'='*60}\n")
      
      # Reiniciar estado
      estado_proceso = {
             "paso_actual": 0,
             "mensaje": "Iniciando evaluación...",
             "porcentaje": 0,
             "completado": False,
             "resultado": None,
             "error": None,
             "usar_excel": usar_excel
      }
      
      # Ejecutar en un hilo separado para no bloquear
      thread = threading.Thread(target=ejecutar_evaluacion)
      thread.daemon = True
      thread.start()
      
      return jsonify({"status": "iniciado"})


@app.route('/api/historial-persona')
def historial_persona():
       """Retorna historial de analisis por DNI o nombre."""
       dni = request.args.get('dni', '').strip()
       nombre = request.args.get('nombre', '').strip()
       if not dni and not nombre:
              return jsonify({'error': 'Debe enviar dni o nombre'}), 400

       if not os.path.exists(HISTORIAL_PERSONAS_PATH):
              return jsonify({'persona': {'dni': dni, 'nombre': nombre}, 'historial': []})

       try:
              with open(HISTORIAL_PERSONAS_PATH, 'r', encoding='utf-8') as f:
                     historial = json.load(f)
              clave = _clave_persona(dni, nombre)
              return jsonify({
                     'persona': {'dni': dni, 'nombre': nombre},
                     'historial': historial.get(clave, [])
              })
       except Exception as e:
              return jsonify({'error': f'No se pudo leer historial: {str(e)}'}), 500

@app.route('/hard-reset', methods=['POST', 'GET'])
def hard_reset():
       """Reinicia forzosamente el estado de la aplicación para desatascarla."""
       global estado_proceso, links_manuales_cti
       links_manuales_cti = []
       estado_proceso = {
              "paso_actual": 0,
              "mensaje": "Listo para evaluar. Haz clic en 'Iniciar Evaluación'.",
              "porcentaje": 0,
              "completado": False,
              "resultado": None,
              "error": None,
              "tiempo_inicio": None,
              "tiempo_estimado_restante": None,
              "cvs_procesados": 0,
              "cvs_total": 0,
              "cvs_exitosos": 0,
              "cvs_con_error": 0,
              "velocidad_cvs_por_segundo": 0,
              "registros_con_error": []
       }
       reset_extraction_progress()
       return jsonify({"status": "ok", "mensaje": "Estado reiniciado correctamente."})

@app.route('/estado')
def obtener_estado():
      """Retorna el estado actual del proceso con información de tiempo real"""
      global estado_proceso
      
      # Obtener progreso de extracción en tiempo real
      progreso_extraccion = get_extraction_progress()
      
      # Calcular tiempo estimado restante
      tiempo_restante = None
      if progreso_extraccion.get("tiempo_inicio") and progreso_extraccion.get("completados", 0) > 0:
             tiempo_transcurrido = time.time() - progreso_extraccion["tiempo_inicio"]
             completados = progreso_extraccion["completados"]
             total = progreso_extraccion["total"]
             
             if completados > 0 and total > completados:
                    velocidad = completados / tiempo_transcurrido  # CVs por segundo
                    restantes = total - completados
                    tiempo_restante = restantes / velocidad if velocidad > 0 else None
                    estado_proceso["velocidad_cvs_por_segundo"] = round(velocidad, 2)
      
      # Actualizar estado con datos de extracción
      estado_proceso["cvs_procesados"] = progreso_extraccion.get("completados", 0)
      estado_proceso["cvs_total"] = progreso_extraccion.get("total", 0)
      estado_proceso["cvs_exitosos"] = progreso_extraccion.get("exitosos", 0)
      estado_proceso["cvs_con_error"] = progreso_extraccion.get("errores", 0)
      estado_proceso["tiempo_estimado_restante"] = round(tiempo_restante, 1) if tiempo_restante else None
      estado_proceso["registros_con_error"] = progreso_extraccion.get("registros_con_error", [])
      
      return jsonify(estado_proceso)


@app.route('/api/reanalizar-registro', methods=['POST'])
def reanalizar_registro():
      """
      Re-analiza un registro individual que tuvo problemas.
      Ãštil para cuando un CV falló en la extracción inicial.
      """
      try:
             data = request.get_json()
             url = data.get('url')
             dni = data.get('dni', '')
             nombre = data.get('nombre', '')
             facultad = data.get('facultad', '')
             carrera = data.get('carrera', '')
             
             if not url:
                    return jsonify({"error": "URL requerida"}), 400
             
             # Crear extractor y re-analizar
             extractor = ExtractorWebCVs()
             resultado = extractor.reanalizar_cv_individual(
                    url, 
                    datos_extra={
                           'dni': dni,
                           'nombre': nombre,
                           'facultad': facultad,
                           'carrera': carrera
                    }
             )
             
             # Si fue exitoso, evaluar el CV
             if resultado['exito'] and resultado['cv_data']:
                    motor = MotorEvaluacion()
                    evaluacion = motor.evaluar_cv_completo(resultado['cv_data'])
                    resultado['evaluacion'] = evaluacion
             
             return jsonify(resultado)
             
      except Exception as e:
             return jsonify({
                    "error": str(e),
                    "exito": False
             }), 500


@app.route('/api/progreso-extraccion')
def obtener_progreso_extraccion():
      """Endpoint dedicado para obtener el progreso de extracción en tiempo real"""
      progreso = get_extraction_progress()
      
      # Calcular estadísticas adicionales
      if progreso.get("tiempo_inicio") and progreso.get("completados", 0) > 0:
             tiempo_transcurrido = time.time() - progreso["tiempo_inicio"]
             completados = progreso["completados"]
             total = progreso["total"]
             
             progreso["tiempo_transcurrido"] = round(tiempo_transcurrido, 1)
             progreso["porcentaje"] = round((completados / total) * 100, 1) if total > 0 else 0
             
             if completados > 0 and total > completados:
                    velocidad = completados / tiempo_transcurrido
                    restantes = total - completados
                    progreso["tiempo_estimado_restante"] = round(restantes / velocidad, 1) if velocidad > 0 else None
                    progreso["velocidad"] = round(velocidad, 2)
             else:
                    progreso["tiempo_estimado_restante"] = 0
                    progreso["velocidad"] = 0
      
      return jsonify(progreso)


@app.route('/api/resumen-pendientes')
def api_resumen_pendientes():
      """
      Retorna el resumen de candidatos con links CTI pendientes agrupados por facultad y carrera.
      Este endpoint puede ser llamado directamente sin ejecutar una evaluación completa.
      """
      try:
             urls, datos_excel, resumen_problemas = leer_datos_desde_requerimientos()
             
             # Preparar datos para el frontend
             total_registros = len(datos_excel) + len(resumen_problemas.get("links_invalidos", []))
             total_sin_link = len(resumen_problemas.get("links_invalidos", []))
             total_con_link = len(datos_excel)
             
             # Ordenar facultades por cantidad de pendientes (mayor a menor)
             facultades_ordenadas = []
             for fac, stats in resumen_problemas.get("resumen_facultades", {}).items():
                    pendientes = stats.get("sin_link", 0) + stats.get("link_invalido", 0)
                    if pendientes > 0:
                           facultades_ordenadas.append({
                                  "facultad": fac,
                                  "total": stats.get("total", 0),
                                  "con_link": stats.get("con_link_valido", 0),
                                  "sin_link": stats.get("sin_link", 0),
                                  "link_invalido": stats.get("link_invalido", 0),
                                  "pendientes": pendientes,
                                  "carreras": stats.get("carreras", {}),
                                  "personas": stats.get("personas_sin_link", []) + stats.get("personas_link_invalido", [])
                           })
             
             facultades_ordenadas.sort(key=lambda x: x["pendientes"], reverse=True)
             
             # Ordenar carreras con problemas
             carreras_ordenadas = []
             for key, stats in resumen_problemas.get("resumen_carreras", {}).items():
                    pendientes = stats.get("sin_link", 0) + stats.get("link_invalido", 0)
                    if pendientes > 0:
                           carreras_ordenadas.append({
                                  "carrera": stats.get("carrera", ""),
                                  "facultad": stats.get("facultad", ""),
                                  "total": stats.get("total", 0),
                                  "con_link": stats.get("con_link_valido", 0),
                                  "pendientes": pendientes,
                                  "personas": stats.get("personas", [])
                           })
             
             carreras_ordenadas.sort(key=lambda x: x["pendientes"], reverse=True)
             
             return jsonify({
                    "success": True,
                    "resumen": {
                           "total_registros": total_registros,
                           "total_con_link": total_con_link,
                           "total_sin_link": total_sin_link,
                           "porcentaje_completado": round((total_con_link / total_registros * 100) if total_registros > 0 else 0, 1)
                    },
                    "facultades": facultades_ordenadas,
                    "carreras": carreras_ordenadas,
                    "links_invalidos": resumen_problemas.get("links_invalidos", [])
             })
             
      except Exception as e:
             print(f"[ERR] Error en api_resumen_pendientes: {e}")
             import traceback
             traceback.print_exc()
             return jsonify({
                    "success": False,
                    "error": str(e)
             }), 500


@app.route('/descargar/<filename>')
def descargar_archivo(filename):
      """Permite descargar reportes generados"""
      return send_from_directory(RESULTADOS_DIR, filename, as_attachment=True)


# â"
# BÃšSQUEDA DE CV EN LEGAJOS (Kit de Contratación)
# â"

def _normalizar_para_match(texto: str) -> str:
      """Normaliza texto para comparación: sin tildes, minúsculas, sin chars especiales, sin espacios extra."""
      import unicodedata, re
      texto = unicodedata.normalize('NFKD', str(texto or ''))
      texto = texto.encode('ascii', 'ignore').decode('ascii')
      texto = re.sub(r'[^a-zA-Z0-9 ]', ' ', texto)
      texto = re.sub(r'\s+', ' ', texto)
      return texto.lower().strip()


def _construir_indice_legajos(carrera_hint: str = '') -> list:
      """
      Construye el índice de carpetas de candidatos usando recorrido de
      PROFUNDIDAD CONTROLADA (máx 4 niveles), SIN os.walk de 50 GB.

      Estructura conocida:
        LEGAJOS_ROOT/
             BP_{nombre}/                                       â† nivel 1
               Kit de Contratación/ (nivel 2 con sub-categorías)
                    00_{CATEGORIA}/                           â† nivel 3
                      {CANDIDATO}/                            â† nivel 4  â† objetivo
               Revisados/ DESESTIMADOS/ (nivel 2 directos)
                    {CANDIDATO}/                                â† nivel 3  â† objetivo

      carrera_hint: si se proporciona, filtra categorías 00_* que no coincidan,
                             reduciendo el tiempo de búsqueda significativamente.

      Retorna lista de dicts: {nombre, nombre_norm, ruta, categoria, bp}
      """
      if not os.path.isdir(LEGAJOS_INPUT_DIR):
             print(f"   [ERR] Dir legajos no encontrado: {LEGAJOS_INPUT_DIR}")
             return []

      # Nivel 2 que actúan como agrupadores de categorías (tienen 00_* dentro)
      L2_CON_CATEGORIAS = {'kit de contrat'}
      # Carpetas de nivel 1 a saltar
      L1_SKIP = {'desuso', 'otros'}
      # Palabras que identifican subcarpetas INTERNAS de un candidato (no es otro candidato)
      INTERNOS = {
             '1 contratac', '2 credencial', '3 experiencia', '4 dp',
             'domiciliado', '3 evaluacion', '2 dni', 'cv', 'docente',
             'no docente', 'otro', '00 doc', 'clases magistrales'
      }

      carrera_norm = _normalizar_para_match(carrera_hint) if carrera_hint else ''
      partes_carrera = [p for p in carrera_norm.split() if len(p) > 3] if carrera_norm else []

      def _es_candidato(nombre_norm: str) -> bool:
             if any(skip in nombre_norm for skip in INTERNOS):
                    return False
             if nombre_norm in L1_SKIP:
                    return False
             palabras = [p for p in nombre_norm.split() if len(p) > 1]
             return len(palabras) >= 2

      indice = []
      print("   [*] Indexando legajos (profundidad controlada)" +
               (f" - filtro: '{carrera_hint}'" if carrera_hint else "") + "...")

      try:
             # â" NIVEL 1: BP_* folders â"
             nivel1 = [
                    d for d in os.listdir(LEGAJOS_INPUT_DIR)
                    if os.path.isdir(os.path.join(LEGAJOS_INPUT_DIR, d))
                    and _normalizar_para_match(d) not in L1_SKIP
             ]

             for d1 in nivel1:
                    ruta1 = os.path.join(LEGAJOS_INPUT_DIR, d1)
                    try:
                           nivel2_dirs = [
                                  d for d in os.listdir(ruta1)
                                  if os.path.isdir(os.path.join(ruta1, d))
                           ]
                    except PermissionError:
                           continue

                    for d2 in nivel2_dirs:
                           ruta2 = os.path.join(ruta1, d2)
                           d2_norm = _normalizar_para_match(d2)

                           # ¿Tiene sub-categorías (Kit de Contratación)?
                           if any(pat in d2_norm for pat in L2_CON_CATEGORIAS):
                                  # NIVEL 3: categorías  00_MEDICINA, 00_CLSâ€¦
                                  try:
                                         nivel3_dirs = [
                                                d for d in os.listdir(ruta2)
                                                if os.path.isdir(os.path.join(ruta2, d))
                                         ]
                                  except PermissionError:
                                         continue

                                  for d3 in nivel3_dirs:
                                         cat_norm = _normalizar_para_match(d3)
                                         ruta3 = os.path.join(ruta2, d3)

                                         # Filtrar por carrera_hint si existe (optimización velocidad)
                                         if partes_carrera:
                                                if not any(p in cat_norm for p in partes_carrera):
                                                       continue  # esta categoría no corresponde a la carrera

                                         # NIVEL 4: candidatos
                                         try:
                                                for d4 in os.listdir(ruta3):
                                                       if not os.path.isdir(os.path.join(ruta3, d4)):
                                                              continue
                                                       d4_norm = _normalizar_para_match(d4)
                                                       if _es_candidato(d4_norm):
                                                              indice.append({
                                                                     'nombre':        d4,
                                                                     'nombre_norm': d4_norm,
                                                                     'ruta':             os.path.join(ruta3, d4),
                                                                     'categoria':   d3,
                                                                     'bp':               d1
                                                              })
                                         except PermissionError:
                                                continue
                           else:
                                  # Nivel 2 directo (Revisados, DESESTIMADOS, etc.) â†’ nivel 3 = candidatos
                                  try:
                                         for d3 in os.listdir(ruta2):
                                                if not os.path.isdir(os.path.join(ruta2, d3)):
                                                       continue
                                                d3_norm = _normalizar_para_match(d3)
                                                if _es_candidato(d3_norm):
                                                       indice.append({
                                                              'nombre':        d3,
                                                              'nombre_norm': d3_norm,
                                                              'ruta':             os.path.join(ruta2, d3),
                                                              'categoria':   d2,
                                                              'bp':               d1
                                                       })
                                  except PermissionError:
                                         continue

      except Exception as e:
             import traceback
             traceback.print_exc()
             print(f"   [!] Error indexando: {e}")

      print(f"      Indice: {len(indice)} candidatos")
      return indice


def buscar_cv_en_legajos(nombre: str, dni: str = '', indice: list = None) -> dict:
      """
      Busca el kit de un candidato en la carpeta de legajos de OneDrive:
        Ingresos o Categorías - Docentes TP - Legajos/
             BP_{BP}/Kit de Contratación/{categoria}/{COD} APELLIDOS, NOMBRES/
               1 CV/               â† CVs principales
               2 CREDENCIALES/  â† títulos y grados
               3 EXPERIENCIA/DOCENTE/ PUBLICACIONES/ NO DOCENTE/ INVESTIGACION/

      Estrategia:
        1. Recorre RECURSIVAMENTE todo el árbol buscando carpetas cuyo nombre
              coincida con el candidato (por nombre o DNI).
        2. Extrae texto de TODOS los PDFs encontrados en las secciones del legajo.
        3. Combina el texto por secciones y lo analiza con ExtractorCV.

      Retorna dict con encontrado, ruta_carpeta, nombre_carpeta, cv_data, evaluacion.
      """
      resultado = {
             "encontrado": False,
             "ruta_carpeta": None,
             "ruta_cv": None,
             "nombre_carpeta": None,
             "cv_data": None,
             "evaluacion": None,
             "candidatos_similares": [],
             "mensaje": ""
      }

      if not os.path.isdir(LEGAJOS_INPUT_DIR):
             resultado["mensaje"] = f"Directorio de legajos no encontrado: {LEGAJOS_INPUT_DIR}"
             print(f"   [ERR] {resultado['mensaje']}")
             return resultado

      nombre_norm = _normalizar_para_match(nombre)
      dni_norm = dni.strip().replace('.', '').replace('-', '').replace(' ', '') if dni else ''
      partes_nombre = [p for p in nombre_norm.split() if len(p) > 2]

      # Usar índice externo (pre-construido) o construir uno nuevo
      entradas = indice if indice is not None else _construir_indice_legajos()

      mejor_match = None
      mejor_score = 0
      todos_candidatos = []

      # Búsqueda en el índice (operación en memoria, O(n) simple)
      for entrada in entradas:
             d = entrada['nombre']
             d_norm = entrada['nombre_norm']
             ruta_cand = entrada['ruta']

             score = 0

             # Match por DNI (máxima prioridad)
             if dni_norm and dni_norm in d_norm.replace(' ', ''):
                    score += 100

             # Match por partes del nombre
             coincidencias_nombre = sum(1 for p in partes_nombre if p in d_norm)
             score += coincidencias_nombre * 20

             # Bonus si nombre completo está contenido
             if nombre_norm and nombre_norm in d_norm:
                    score += 50

             if score >= 40:
                    todos_candidatos.append({'nombre': d, 'ruta': ruta_cand, 'score': score})
                    if score > mejor_score:
                           mejor_score = score
                           mejor_match = {'nombre': d, 'ruta': ruta_cand}

      # Guardar similares para el frontend
      resultado["candidatos_similares"] = sorted(
             todos_candidatos, key=lambda x: x["score"], reverse=True
      )[:5]

      if not mejor_match or mejor_score < 40:
             resultado["mensaje"] = (
                    f"No se encontró carpeta de legajo para '{nombre}'"
                    + (f" (DNI: {dni})" if dni else "")
                    + f". Buscado en: {LEGAJOS_INPUT_DIR}"
             )
             print(f"   [ERR] {resultado['mensaje']}")
             return resultado

      ruta_candidato = mejor_match["ruta"]
      resultado["ruta_carpeta"] = ruta_candidato
      resultado["nombre_carpeta"] = mejor_match["nombre"]
      resultado["encontrado"] = True
      print(f"   [OK] Carpeta encontrada (score={mejor_score}): {mejor_match['nombre']}")

      # â" Recopilar TODOS los PDFs del legajo por sección â"
      def _recopilar_pdfs_seccion(ruta_base: str, subfolder: str):
             """Retorna lista de rutas PDF dentro de una subcarpeta (si existe)."""
             ruta = os.path.join(ruta_base, subfolder)
             if not os.path.isdir(ruta):
                    return []
             pdfs = []
             for rroot, _, rfiles in os.walk(ruta):
                    for f in rfiles:
                           if f.lower().endswith('.pdf'):
                                  pdfs.append(os.path.join(rroot, f))
             return pdfs

      def _extraer_texto_pdf(ruta_pdf: str) -> str:
             """Extrae texto de un PDF usando pdfplumber o PyPDF2 como fallback."""
             texto = ''
             try:
                    import pdfplumber
                    with pdfplumber.open(ruta_pdf) as pdf:
                           for page in pdf.pages:
                                  t = page.extract_text() or ''
                                  texto += t + '\n'
             except Exception:
                    try:
                           import PyPDF2
                           with open(ruta_pdf, 'rb') as f:
                                  reader = PyPDF2.PdfReader(f)
                                  for page in reader.pages:
                                         texto += (page.extract_text() or '') + '\n'
                    except Exception:
                           pass
             return texto.strip()

      # â" Recolectar TODOS los PDFs del kit por subcarpeta (estructura agnóstica) â"
      # Funciona con CUALQUIER estructura de carpetas:
      #   Kit de Contratación: 1 CONTRATACION/CV, 2 CREDENCIALES, 3 EXPERIENCIA, 4 DP
      #   Revisados:  1. INFORMACIÃ\"N DE CONTRATACIÃ\"N, 2. FORMACIÃ\"N ACADÉMICA, etc.
      #
      # Estrategia: recorrer TODAS las subcarpetas del candidato con os.walk,
      # agrupar PDFs por subcarpeta inmediata (nivel 1) para el label,
      # y asignar un label semántico según palabras clave en el nombre.

      # Mapa de palabras clave â†’ label semántico
      _KWORD_LABEL = [
             (['contrat', 'cv', '1 contrat', '1. info'],             "=== CV / CONTRATACIÃ\"N ==="),
             (['credencial', 'formac', 'academ', 'titulo', '2 cred', '2. form'], "=== CREDENCIALES Y FORMACIÃ\"N ==="),
             (['experiencia', 'laboral', 'docente', 'no docente', '3 exp', '3. exp'], "=== EXPERIENCIA ==="),
             (['dp', 'descripcion', 'puesto', '4 dp', '5. desc'],   "=== DESCRIPCIÃ\"N DE PUESTO ==="),
             (['capacit', 'desarrollo', '4. des'],                   "=== CAPACITACIONES ==="),
      ]

      def _label_para_carpeta(nombre_carpeta: str) -> str:
             n = _normalizar_para_match(nombre_carpeta)
             for palabras, label in _KWORD_LABEL:
                    if any(p in n for p in palabras):
                           return label
             return f"=== {nombre_carpeta.upper()} ==="

      bloques_texto   = []
      pdfs_encontrados = []
      pdfs_vistos      = set()   # evitar duplicados por symlinks / rutas solapadas

      # Paso 1: iterar subcarpetas de nivel 1 del candidato (orden alfanumérico â†’ sigue numeración)
      try:
             nivel1 = sorted([
                    d for d in os.listdir(ruta_candidato)
                    if os.path.isdir(os.path.join(ruta_candidato, d))
             ])
      except Exception:
             nivel1 = []

      for sub1 in nivel1:
             ruta_sub1   = os.path.join(ruta_candidato, sub1)
             label_sub1  = _label_para_carpeta(sub1)
             pdfs_sub1   = []
             for rroot, _, rfiles in os.walk(ruta_sub1):
                    for f in rfiles:
                           if f.lower().endswith('.pdf'):
                                  ruta_pdf = os.path.join(rroot, f)
                                  if ruta_pdf not in pdfs_vistos:
                                         pdfs_vistos.add(ruta_pdf)
                                         pdfs_sub1.append(ruta_pdf)
             if pdfs_sub1:
                    pdfs_encontrados.extend(pdfs_sub1)
                    textos_sub = []
                    for pdf_path in pdfs_sub1:
                           t = _extraer_texto_pdf(pdf_path)
                           if t:
                                  textos_sub.append(f"[{os.path.basename(pdf_path)}]\n{t}")
                    if textos_sub:
                           bloques_texto.append(f"\n{label_sub1}\n" + "\n\n".join(textos_sub))

      # Paso 2: PDFs directamente en la raíz (sin subcarpeta)
      try:
             for f in os.listdir(ruta_candidato):
                    if f.lower().endswith('.pdf'):
                           ruta_pdf = os.path.join(ruta_candidato, f)
                           if ruta_pdf not in pdfs_vistos:
                                  pdfs_vistos.add(ruta_pdf)
                                  pdfs_encontrados.append(ruta_pdf)
                                  t = _extraer_texto_pdf(ruta_pdf)
                                  if t:
                                         bloques_texto.append(f"[{f}]\n{t}")
      except Exception:
             pass

      if not pdfs_encontrados:
             resultado["mensaje"] = f"Carpeta '{mejor_match['nombre']}' encontrada pero sin PDFs"
             print(f"   [!] {resultado['mensaje']}")
             return resultado

      pass  # log removed (encoding issue)
      resultado["ruta_cv"] = pdfs_encontrados[0]  # primer PDF como referencia
      resultado["archivo_cv"] = f"{len(pdfs_encontrados)} archivos PDF"

      texto_combinado = "\n\n".join(bloques_texto)
      if not texto_combinado.strip():
             resultado["mensaje"] = f"PDFs encontrados ({len(pdfs_encontrados)}) pero sin texto extraíble (imágenes escaneadas)"
             print(f"   [!] {resultado['mensaje']}")
             return resultado

      # â" Analizar texto combinado con ExtractorCV â"
      try:
             from extractor_cvs import ExtractorCV
             # Crear extractor con ruta válida pero sobreescribir el texto
             extractor = ExtractorCV(pdfs_encontrados[0])
             # Inyectar texto combinado (evita que extraer_texto() sobreescriba)
             extractor.texto_completo = texto_combinado

             try:      educacion  = extractor.extraer_nivel_educacion()
             except Exception: educacion = {"doctorado": False, "maestria": False, "licenciatura": False, "bachiller": False}
             try:      anos_exp   = extractor.extraer_anos_experiencia()
             except Exception: anos_exp = 0
             try:      publicaciones = extractor.extraer_publicaciones()
             except Exception: publicaciones = 0
             try:      nombre_ext = extractor.extraer_nombre()
             except Exception: nombre_ext = nombre

             # Experiencia docente — análisis estructurado de periodos en el texto
             # (cubre CVs extensos donde la docencia aparece como rangos de fechas
             #  y no como la frase "X años de experiencia docente")
             import re as _re
             exp_docente = 0
             experiencia_laboral_detectada = []
             evidencias_docencia = []
             try:
                    from analizador_experiencia import analizar_experiencia_texto
                    analisis_exp = analizar_experiencia_texto(texto_combinado)
                    exp_docente = analisis_exp["anos_docencia"]
                    experiencia_laboral_detectada = analisis_exp["experiencia_laboral"]
                    evidencias_docencia = analisis_exp["evidencias_docencia"]
                    if analisis_exp["anos_totales"] > anos_exp:
                           anos_exp = analisis_exp["anos_totales"]
             except Exception as e_exp:
                    print(f"   [!] analizador_experiencia fallo: {e_exp}")
             # Fallback: frase explícita "X años de experiencia docente"
             match_doc = _re.search(r'(\d+)\s*a[ñn]os?\s+(?:de\s+)?(?:experiencia\s+)?docente', texto_combinado, _re.IGNORECASE)
             if match_doc:
                    exp_docente = max(exp_docente, int(match_doc.group(1)))

             # Idiomas
             idiomas = []
             try:
                    for idioma in ['ingles', 'frances', 'aleman', 'portugues', 'italiano', 'chino', 'japones']:
                           if idioma in _normalizar_para_match(texto_combinado):
                                  idiomas.append(idioma.capitalize())
             except Exception:
                    pass

             cv_data = {
                    "nombre":                      nombre_ext or nombre,
                    "dni":                            dni,
                    "texto_completo":        texto_combinado,
                    "fuente":                      "legajo_kit",
                    "fuente_datos":             "LEGAJO_PDF",
                    "archivo":                     f"{len(pdfs_encontrados)} PDFs del legajo",
                    "educacion":                  educacion,
                    "anos_experiencia":      anos_exp,
                    "experiencia_docente": exp_docente,
                    "publicaciones":           publicaciones,
                    "proyectos_investigacion": 0,
                    "idiomas":                     idiomas,
                    "experiencia_laboral": experiencia_laboral_detectada,
                    "evidencias_docencia": evidencias_docencia,
                    "premios":                     [],
                    "url":                            "",
             }

             cv_data["nombre"] = cv_data["nombre"] or nombre
             cv_data["dni"]      = cv_data["dni"] or dni

             resultado["cv_data"] = cv_data

             # â" Evaluar con motor de rúbrica â"
             motor = MotorEvaluacion()
             evaluacion = motor.evaluar_cv_completo(cv_data)
             evaluacion["fuente_datos"]      = "LEGAJO_PDF"
             evaluacion["carpeta_legajo"]  = mejor_match["nombre"]
             evaluacion["pdfs_analizados"] = len(pdfs_encontrados)

             try:
                    from generador_decisiones_mejorado import GeneradorDecisiones
                    gen = GeneradorDecisiones([evaluacion])
                    gen.clasificar_todos()
                    if gen.clasificaciones:
                           cl = gen.clasificaciones[0]
                           evaluacion["clasificacion"] = cl.get("perfil", evaluacion.get("clasificacion", ""))
                           evaluacion["es_elegible"]   = cl.get("elegible", False)
                           evaluacion["justificacion_decision"] = cl.get("justificacion", "")
             except Exception as e_dec:
                    print(f"   [!] No se pudo clasificar: {e_dec}")

             resultado["evaluacion"] = evaluacion
             resultado["mensaje"] = (
                    f"{len(pdfs_encontrados)} PDFs analizados de '{mejor_match['nombre']}'"
             )
             pass  # log removed (encoding issue)
                      # continuation removed

      except Exception as e:
             import traceback
             traceback.print_exc()
             resultado["mensaje"] += f" — Error al analizar: {str(e)}"

      return resultado


@app.route('/api/buscar-cv-legajo', methods=['POST'])
def api_buscar_cv_legajo():
      """
      Busca el CV físico de un candidato en los legajos del kit de contratación
      (SISTEMA DE VALIDACION DE PERSONAL/INPUT/) y realiza la evaluación por rúbrica.

      Body JSON: { nombre, dni (opcional), facultad (opcional), carrera (opcional) }
      """
      try:
             data = request.get_json() or {}
             nombre = data.get('nombre', '').strip()
             dni = data.get('dni', '').strip()
             facultad = data.get('facultad', '')
             carrera = data.get('carrera', '')

             if not nombre and not dni:
                    return jsonify({"error": "Se requiere al menos nombre o DNI"}), 400

             print(f"\n{'='*60}")
             pass  # log removed (encoding issue)
             print(f"   Nombre: {nombre}")
             print(f"   DNI: {dni}")
             print(f"   Directorio: {LEGAJOS_INPUT_DIR}")
             print(f"{'='*60}")

             resultado = buscar_cv_en_legajos(nombre, dni)

             # Enriquecer evaluacion con datos de facultad/carrera
             if resultado.get("evaluacion"):
                    resultado["evaluacion"]["facultad"] = resultado["evaluacion"].get("facultad") or facultad
                    resultado["evaluacion"]["carrera"] = resultado["evaluacion"].get("carrera") or carrera

             print(f"   -> Encontrado: {resultado['encontrado']}")
             print(f"   -> Mensaje: {resultado['mensaje']}")
             if resultado.get("evaluacion"):
                    ev = resultado["evaluacion"]
                    print(f"   -> Puntaje: {ev.get('total', 0)} | Perfil: {ev.get('clasificacion', 'N/A')}")

             return jsonify({
                    "exito": resultado["encontrado"] and resultado.get("evaluacion") is not None,
                    "encontrado": resultado["encontrado"],
                    "nombre_carpeta": resultado.get("nombre_carpeta"),
                    "archivo_cv": os.path.basename(resultado["ruta_cv"]) if resultado.get("ruta_cv") else None,
                    "mensaje": resultado["mensaje"],
                    "candidatos_similares": resultado.get("candidatos_similares", []),
                    "evaluacion": resultado.get("evaluacion"),
                    "cv_data": {
                           k: v for k, v in (resultado.get("cv_data") or {}).items()
                           if k not in ("texto_completo",)  # no enviar texto crudo al frontend
                    } if resultado.get("cv_data") else None
             })

      except Exception as e:
             import traceback
             traceback.print_exc()
             return jsonify({"error": str(e), "exito": False}), 500


# â"
#  MOTOR DE INFERENCIA PROFESIONAL PARA CV RECONSTRUIDOS DESDE LA WEB
# â"
def _inferir_campos_cv_web(cv_web: dict) -> None:
    """
    Enriquece los campos de un cv_data reconstruido desde la web aplicando
    inferencia profesional basada en rol + institución + años de experiencia.

    Reglas de inferencia:
      1. experiencia_laboral vacía  â†’ se construye desde cargo_actual e instituciones
      2. empresas vacías            â†’ se poblan desde instituciones detectadas
      3. experiencia_docente == 0   â†’ se estima si hay cargo directivo en institución
                                      médica/académica (docencia tutoral implícita)
      4. texto_completo vacío/corto â†’ se enriquece con cargos e instituciones para
                                      potenciar el keyword-matching del motor
    """
    import re as _re

    instituciones   = cv_web.get('instituciones', [])
    cargo_actual    = cv_web.get('cargo_actual', '')
    empresa_actual  = cv_web.get('empresa_actual', '')
    anos_exp        = int(cv_web.get('anos_experiencia', 0) or 0)

    # â" 1. Construir experiencia_laboral â"
    exp_lab = cv_web.get('experiencia_laboral', [])
    if not exp_lab:
        if cargo_actual:
            exp_lab.append({
                'cargo':   cargo_actual,
                'empresa': empresa_actual or (instituciones[0] if instituciones else ''),
                'anos':    anos_exp,
            })
        for inst in instituciones[:3]:
            if inst and not any(e.get('empresa', '') == inst for e in exp_lab):
                exp_lab.append({'cargo': '', 'empresa': inst, 'anos': 0})
        cv_web['experiencia_laboral'] = exp_lab

    # â" 2. Construir lista de empresas â"
    if not cv_web.get('empresas'):
        empresas_raw = [empresa_actual] + instituciones if empresa_actual else list(instituciones)
        cv_web['empresas'] = list(dict.fromkeys(e for e in empresas_raw if e))[:6]

    # â" 3. Enriquecer texto_completo con cargos, instituciones y snippets â"
    texto_actual = cv_web.get('texto_completo', '') or ''
    extras_parts = [cargo_actual, empresa_actual] + instituciones
    # Incluir snippets de directorios (pueden tener cargo + institución completos)
    for sn in cv_web.get('snippet_directorio', []):
        if sn:
            extras_parts.append(sn[:300])
    extras = ' '.join(filter(None, extras_parts))
    if extras and extras.lower() not in texto_actual.lower():
        cv_web['texto_completo'] = (texto_actual + ' ' + extras).strip()

    # â" 4. Inferencia: experiencia_docente desde rol + institución â"
    exp_doc_actual = cv_web.get('experiencia_docente', 0)
    if isinstance(exp_doc_actual, list):
        exp_doc_actual = 0  # ya será procesada por la normalización del motor
    exp_doc_actual = int(exp_doc_actual or 0)

    # Extraer cargo desde texto cuando cargo_actual está vacío
    if not cargo_actual:
        texto_base = cv_web.get('texto_combinado', '') or cv_web.get('texto_completo', '')
        m_cargo = _re.search(
            r'\b(?:director\s+(?:del?\s+)?(?:departamento|servicio|ejecutivo|m[eé]dico|general)'
            r'|jefe\s+(?:del?\s+)?(?:departamento|servicio|unidad)'
            r'|m[eé]dico\s+(?:jefe|especialista)'
            r'|coordinador\s+m[eé]dico)',
            texto_base, _re.I)
        if m_cargo:
            cargo_actual = m_cargo.group(0)
            cv_web.setdefault('cargo_actual', cargo_actual)

    if exp_doc_actual == 0 and anos_exp >= 3:
        texto_inferencia = (
            cargo_actual + ' ' +
            empresa_actual + ' ' +
            ' '.join(instituciones) + ' ' +
            ' '.join(cv_web.get('snippet_directorio', [])) + ' ' +
            cv_web.get('texto_combinado', '')
        ).lower()

        _DOCENTE_LITERAL   = ['docente', 'profesor', 'catedrático', 'catedratico', 'instructor']
        _CARGOS_DIRECTIVOS = [
            'director de', 'director del', 'director médico', 'director medico',
            'jefe de', 'jefe del', 'jefatura', 'coordinador',
            'decano', 'subdirector', 'gerente', 'rector', 'vicerrector',
            'médico jefe', 'medico jefe', 'presidente de',
        ]
        _INST_ACADEMICAS   = [
            'hospital', 'clínica', 'clinica', 'inen', 'instituto',
            'ministerio', 'essalud', 'universidad', 'escuela', 'facultad',
            'oncosalud', 'ins.gob', 'minsa', 'ins ', 'solidaridad',
        ]

        tiene_docente    = any(x in texto_inferencia for x in _DOCENTE_LITERAL)
        tiene_directivo  = any(x in texto_inferencia for x in _CARGOS_DIRECTIVOS)
        tiene_inst_acad  = any(x in texto_inferencia for x in _INST_ACADEMICAS)

        if tiene_docente:
            # Cargo docente literal detectado â†’ estimación conservadora proporcional
            cv_web['experiencia_docente'] = max(1, min(15, int(anos_exp * 0.4)))
        elif tiene_directivo and tiene_inst_acad and anos_exp >= 5:
            # Directivo en institución médica/académica â†’ docencia tutoral implícita
            cv_web['experiencia_docente'] = max(2, min(10, int(anos_exp * 0.3)))
        elif tiene_directivo and anos_exp >= 12:
            # Senior sin institución clara â†’ mínima docencia plausible
            cv_web['experiencia_docente'] = 3

        if cv_web.get('experiencia_docente', 0):
            pass  # log removed (encoding issue)
                  # continuation removed

    # â" 5. Boost de confianza cuando la inferencia detectó perfil profesional â"
    conf_actual = cv_web.get('score_confianza', 0)
    boost = 0
    if cargo_actual:                          boost += 10
    if empresa_actual or instituciones:       boost += 10
    if cv_web.get('experiencia_docente', 0):  boost += 5
    if anos_exp >= 10:                        boost += 10
    if boost > 0:
        cv_web['score_confianza'] = min(100, conf_actual + boost)
        pass  # log removed (encoding issue)


@app.route('/api/buscar-masivo-nombres', methods=['POST'])
def api_buscar_masivo_nombres():
      """
      Búsqueda masiva de candidatos por lista de nombres pegada.
      Body JSON: { "nombres": ["Juan Pérez","María López",...] }
      """
      global estado_proceso
      data    = request.get_json() or {}
      nombres = [n.strip() for n in data.get('nombres', []) if str(n).strip()]

      if not nombres:
             return jsonify({"error": "Lista de nombres vacía"}), 400

      # Reiniciar estado de proceso
      estado_proceso.update({
             "paso_actual":               0,
             "mensaje":                   f"Iniciando búsqueda masiva de {len(nombres)} candidatos...",
             "porcentaje":                0,
             "completado":                False,
             "resultado":                 None,
             "error":                     None,
             "tiempo_inicio":             time.time(),
             "tiempo_estimado_restante":  None,
             "cvs_procesados":            0,
             "cvs_total":                 len(nombres),
             "cvs_exitosos":              0,
             "cvs_con_error":             0,
             "velocidad_cvs_por_segundo": 0,
             "registros_con_error":       []
      })

      hilo = threading.Thread(
             target=_ejecutar_busqueda_masiva_nombres,
             args=(nombres,),
             daemon=True
      )
      hilo.start()
      return jsonify({"status": "iniciado"})


def _ejecutar_busqueda_masiva_nombres(nombres: list):
      """Hilo de fondo: busca y evalúa cada nombre en la web."""
      global estado_proceso
      try:
             from buscador_web_cv import buscar_cv_en_web
      except ImportError:
             estado_proceso.update({
                    "completado": True,
                    "error": "Módulo buscador_web_cv no disponible. Instala: pip install requests beautifulsoup4"
             })
             return

      evaluaciones = []
      total        = len(nombres)

      for i, nombre in enumerate(nombres):
             pct = int(((i) / total) * 80) + 10
             estado_proceso.update({
                    "paso_actual":    2,
                    "mensaje":        f" Buscando [{i+1}/{total}]: {nombre[:45]}...",
                    "porcentaje":     pct,
                    "cvs_procesados": i
             })
             try:
                    resultado_web = buscar_cv_en_web(nombre, '', '', '')
                    if resultado_web.get("encontrado") and resultado_web.get("cv_data"):
                           cv_web = resultado_web["cv_data"]
                           cv_web["nombre"]       = nombre
                           cv_web["fuente_datos"] = "WEB_MASIVO"
                           cv_web["fuente"]       = "web"
                           cv_web["url"]          = (cv_web.get("fuentes_urls") or [""])[0]
                           if "texto_combinado" in cv_web and "texto_completo" not in cv_web:
                                  cv_web["texto_completo"] = cv_web["texto_combinado"]
                           cv_web.setdefault("experiencia_laboral", [])
                           cv_web.setdefault("premios", [])
                           _inferir_campos_cv_web(cv_web)

                           motor    = MotorEvaluacion()
                           evaluacion = motor.evaluar_cv_completo(cv_web)
                           evaluacion["fuente_datos"]        = "WEB_MASIVO"
                           evaluacion["reconstruido_web"]    = True
                           evaluacion["score_confianza_web"] = cv_web.get("score_confianza", 0)
                           evaluacion["fuentes_urls"]        = cv_web.get("fuentes_urls", [])
                           estado_proceso["cvs_exitosos"] = estado_proceso.get("cvs_exitosos", 0) + 1
                    else:
                           evaluacion = {
                                  "nombre":          nombre, "dni": "",
                                  "total":           0, "maximo": 200, "porcentaje": 0.0,
                                  "clasificacion":   "NO_ELEGIBLE", "es_elegible": False,
                                  "tipo_perfil":     "general",
                                  "puntajes":        {"C1":0,"C2":0,"C3":0,"C4":0,"C5":0,"C6":0,"C7":0},
                                  "detalles":        {}, "archivo": "", "fuente": "web",
                                  "fuente_datos":    "WEB_MASIVO",
                                  "error_extraccion": True,
                                  "error_mensaje":   "No encontrado en la web"
                           }
                           estado_proceso["registros_con_error"].append({
                                  "nombre": nombre, "error": "No encontrado en la web"
                           })
                           estado_proceso["cvs_con_error"] = estado_proceso.get("cvs_con_error", 0) + 1
             except Exception as exc:
                    evaluacion = {
                           "nombre":          nombre, "dni": "",
                           "total":           0, "maximo": 200, "porcentaje": 0.0,
                           "clasificacion":   "NO_ELEGIBLE", "es_elegible": False,
                           "tipo_perfil":     "general",
                           "puntajes":        {"C1":0,"C2":0,"C3":0,"C4":0,"C5":0,"C6":0,"C7":0},
                           "detalles":        {}, "archivo": "", "fuente": "web",
                           "fuente_datos":    "WEB_MASIVO",
                           "error_extraccion": True,
                           "error_mensaje":   str(exc)
                    }
                    estado_proceso["registros_con_error"].append({
                           "nombre": nombre, "error": str(exc)
                    })
                    estado_proceso["cvs_con_error"] = estado_proceso.get("cvs_con_error", 0) + 1

             evaluaciones.append(evaluacion)

      # Construir resultado en el mismo formato que /iniciar-evaluacion
      resultado = {
             "evaluaciones": [
                    {
                           "nombre":              e.get("nombre", ""),
                           "dni":                 e.get("dni", ""),
                           "total":               e.get("total", 0),
                           "maximo":              e.get("maximo", 200),
                           "porcentaje":          e.get("porcentaje", 0.0),
                           "clasificacion":       e.get("clasificacion", ""),
                           "es_elegible":         e.get("es_elegible", False),
                           "tipo_perfil":         e.get("tipo_perfil", "general"),
                           "puntajes":            e.get("puntajes", {}),
                           "detalles":            e.get("detalles", {}),
                           "archivo":             e.get("archivo", ""),
                           "fuente":              e.get("fuente", "web"),
                           "fuente_datos":        e.get("fuente_datos", "WEB_MASIVO"),
                           "facultad":            e.get("facultad", ""),
                           "carrera":             e.get("carrera", ""),
                           "url":                 e.get("url", ""),
                           "error_extraccion":    e.get("error_extraccion", False),
                           "error_mensaje":       e.get("error_mensaje", ""),
                           "perfil_desactualizado": False,
                           "perfil_vacio":        False,
                           "meses_sin_actualizar": None,
                           "fecha_actualizacion": None,
                           "info_incompleta":     e.get("error_extraccion", False),
                           "justificacion_decision": e.get("justificacion_decision", ""),
                           "reconstruido_web":    e.get("reconstruido_web", False),
                           "score_confianza_web": e.get("score_confianza_web", 0)
                    }
                    for e in evaluaciones
             ],
             "reportes":             {"json": "busqueda_masiva.json"},
             "fecha":                datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
             "resumen_problemas":    {"links_invalidos": [], "resumen_facultades": {}},
             "perfiles_con_problemas": []
      }

      estado_proceso.update({
             "paso_actual":    5,
             "mensaje":        f" Búsqueda masiva completada — {total} candidatos procesados",
             "porcentaje":     100,
             "cvs_procesados": total,
             "completado":     True,
             "resultado":      resultado
      })


@app.route('/api/buscar-cv-web', methods=['POST'])
def api_buscar_cv_web():
      """
      Búsqueda EXPERIMENTAL de CV en la web (LinkedIn, CTI Vitae, páginas públicas).
      Body JSON: { nombre, dni (opcional), facultad (opcional), carrera (opcional) }
      """
      try:
             data     = request.get_json() or {}
             nombre   = data.get('nombre',   '').strip()
             dni      = data.get('dni',      '').strip()
             facultad = data.get('facultad', '').strip()
             carrera  = data.get('carrera',  '').strip()

             if not nombre:
                    return jsonify({"error": "Se requiere el nombre del candidato"}), 400

             try:
                    from buscador_web_cv import buscar_cv_en_web
             except ImportError:
                    return jsonify({"error": "Módulo buscador_web_cv no disponible. Instala: pip install requests beautifulsoup4"}), 500

             print(f"\n{'='*60}")
             pass  # log removed (encoding issue)
             print(f"{'='*60}")
             resultado = buscar_cv_en_web(nombre, dni, facultad, carrera)

             if not resultado.get("encontrado") or not resultado.get("cv_data"):
                    return jsonify({
                           "exito": False,
                           "encontrado": False,
                           "mensaje": resultado.get("mensaje", "No se encontró información en la web."),
                    })

             # â" Normalizar cv_data para que sea compatible con MotorEvaluacion â"
             cv_web = resultado["cv_data"]
             cv_web["nombre"]         = nombre
             cv_web["dni"]            = dni or cv_web.get("dni", "")
             cv_web["facultad"]       = facultad
             cv_web["carrera"]        = carrera
             cv_web["fuente_datos"]   = "WEB_EXPERIMENTAL"
             cv_web["fuente"]         = "web"
             cv_web["url"]            = (cv_web.get("fuentes_urls") or [""])[0]
             # Alias para MotorEvaluacion (espera 'texto_completo')
             if "texto_combinado" in cv_web and "texto_completo" not in cv_web:
                    cv_web["texto_completo"] = cv_web["texto_combinado"]
             # Asegurar campos mínimos
             for campo in ["experiencia_laboral", "premios"]:
                    cv_web.setdefault(campo, [])

             # â" Inferencia profesional post-enriquecimiento â"
             # Ejecutar DESPUÃ‰S de que directorios/RENATI/OpenAlex ya enriquecieron cv_web
             _inferir_campos_cv_web(cv_web)

             # â" Log diagnóstico (siempre visible en consola) â"
             pass  # log removed (encoding issue)
             pass  # log removed (encoding issue)
             print(f"   [CV_WEB] instituciones={cv_web.get('instituciones',[])[:3]}")
             print(f"   [CV_WEB] empresas={cv_web.get('empresas',[])[:3]}")
             print(f"   [CV_WEB] anos_exp={cv_web.get('anos_experiencia',0)}  exp_docente={cv_web.get('experiencia_docente',0)}")
             print(f"   [CV_WEB] educacion={cv_web.get('educacion',{}).get('grado_maximo','?')}")
             print(f"   [CV_WEB] confianza={cv_web.get('score_confianza',0)}%")

             # â" Evaluar con rúbrica â"
             evaluacion = {}
             try:
                    motor_web = MotorEvaluacion()
                    evaluacion = motor_web.evaluar_cv_completo(cv_web)
                    evaluacion["fuente_datos"]      = "WEB_EXPERIMENTAL"
                    evaluacion["nombre"]            = nombre
                    evaluacion["dni"]               = dni or cv_web.get("dni", "")
                    evaluacion["facultad"]          = facultad
                    evaluacion["carrera"]           = carrera
                    evaluacion["reconstruido_web"]  = True
                    evaluacion["score_confianza_web"] = cv_web.get("score_confianza", 0)
                    evaluacion["fuentes_urls"]      = cv_web.get("fuentes_urls", [])

                    # Clasificar con GeneradorDecisiones
                    try:
                           from generador_decisiones_mejorado import GeneradorDecisiones
                           gen = GeneradorDecisiones([evaluacion])
                           gen.clasificar_todos()
                           if gen.clasificaciones:
                                  cl = gen.clasificaciones[0]
                                  evaluacion["clasificacion"]           = cl.get("perfil", evaluacion.get("clasificacion", ""))
                                  evaluacion["es_elegible"]             = cl.get("elegible", False)
                                  evaluacion["justificacion_decision"]  = cl.get("justificacion", "")
                    except Exception as e_dec:
                           pass  # log removed (encoding issue)

                    print(f"   [OK] Web clasificado: {evaluacion.get('clasificacion','?')} | "
                             f"Puntaje: {evaluacion.get('total','?')}/200 | "
                             f"Confianza web: {cv_web.get('score_confianza',0)}%")
             except Exception as e_motor:
                    print(f"   [!] Error en MotorEvaluacion (web): {e_motor}")

             # Persistir resultado en estado_proceso para que Excel coincida
             try:
                    if estado_proceso.get("resultado") and evaluacion:
                           for i, ev in enumerate(estado_proceso["resultado"].get("evaluaciones", [])):
                                  match_dni = dni and str(ev.get("dni", "")).strip() == str(dni).strip()
                                  match_nom = nombre and ev.get("nombre", "").lower().strip() == str(nombre).lower().strip()
                                  if match_dni or match_nom:
                                         estado_proceso["resultado"]["evaluaciones"][i].update({
                                                "total":                  evaluacion.get("total", ev.get("total")),
                                                "porcentaje":             evaluacion.get("porcentaje", ev.get("porcentaje")),
                                                "clasificacion":          evaluacion.get("clasificacion", ev.get("clasificacion")),
                                                "es_elegible":            evaluacion.get("es_elegible", ev.get("es_elegible")),
                                                "tipo_perfil":            evaluacion.get("tipo_perfil", ev.get("tipo_perfil", "general")),
                                                "puntajes":               evaluacion.get("puntajes", ev.get("puntajes")),
                                                "detalles":               evaluacion.get("detalles", ev.get("detalles", {})),
                                                "justificacion_decision": evaluacion.get("justificacion_decision", ""),
                                                "reconstruido_web":       True,
                                                "fuente_datos":           "WEB_EXPERIMENTAL",
                                         })
                                         print(f"   [OK] estado_proceso actualizado para DNI={dni} nombre={nombre}")
                                         break
             except Exception as e_persist:
                    print(f"   [!] No se pudo persistir en estado_proceso: {e_persist}")

             return jsonify({
                    "exito":           True,
                    "encontrado":      True,
                    "mensaje":         resultado.get("mensaje", ""),
                    "evaluacion":      evaluacion,
                    "score_confianza": cv_web.get("score_confianza", 0),
                    "fuentes_urls":    cv_web.get("fuentes_urls", []),
                    "cv_data": {
                           k: v for k, v in cv_web.items()
                           if k not in ("texto_completo", "texto_combinado", "snippets")
                    },
             })

      except Exception as e:
             import traceback
             traceback.print_exc()
             return jsonify({"error": str(e), "exito": False}), 500




@app.route('/descargar-analisis-links')
def descargar_analisis_links():
      """Genera y descarga Excel con análisis completo de links CTI (columna J)"""
      global archivo_excel_subido
      
      try:
             import openpyxl
             from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
             
             # Usar el archivo Excel subido por el usuario
             if archivo_excel_subido and os.path.exists(archivo_excel_subido):
                    archivo_fuente = archivo_excel_subido
             else:
                    # Fallback: buscar en ubicaciones conocidas
                    posibles_rutas = [
                           os.path.join(os.path.dirname(sys.executable), '..', 'Extraccion de data', 'Requerimiento docentes 2026-1 300126.xlsx'),
                           os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Extraccion de data', 'Requerimiento docentes 2026-1 300126.xlsx'),
                    ]
                    archivo_fuente = None
                    for ruta in posibles_rutas:
                           if os.path.exists(ruta):
                                  archivo_fuente = ruta
                                  break
                    
                    if not archivo_fuente:
                           return jsonify({"error": "Por favor, sube primero un archivo Excel con los datos de requerimientos"}), 400
             
             # Leer archivo fuente
             wb_fuente = openpyxl.load_workbook(archivo_fuente)
             ws_fuente = wb_fuente.active
             
             # Crear nuevo workbook para el reporte
             wb_reporte = openpyxl.Workbook()
             
             # Estilos
             titulo_font = Font(bold=True, size=14, color="FFFFFF")
             header_font = Font(bold=True, size=11)
             titulo_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
             verde_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
             rojo_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
             amarillo_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
             gris_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
             borde = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
             )
             
             # HOJA 1: RESUMEN
             ws_resumen = wb_reporte.active
             ws_resumen.title = "RESUMEN"
             
             ws_resumen.merge_cells('A1:D1')
             ws_resumen['A1'] = "REPORTE DE ANALISIS - LINKS CTI (COLUMNA J)"
             ws_resumen['A1'].font = titulo_font
             ws_resumen['A1'].fill = titulo_fill
             ws_resumen['A1'].alignment = Alignment(horizontal='center')
             
             ws_resumen['A2'] = f"Fecha de generacion: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
             ws_resumen['A3'] = f"Archivo analizado: Requerimiento docentes 2026-1 300126.xlsx"
             
             # Conteos
             conteo = {'links_validos': 0, 'en_construccion': 0, 'no_tiene': 0, 'vacio': 0, 'otro': 0}
             resultados = []
             
             for row in range(2, 103):
                    nombre = ws_fuente.cell(row=row, column=2).value
                    col_a = ws_fuente.cell(row=row, column=1).value
                    valor_j = ws_fuente.cell(row=row, column=10).value
                    
                    if valor_j is None or str(valor_j).strip() == "":
                           estado = "VACIO (Sin datos)"
                           conteo['vacio'] += 1
                    elif str(valor_j).strip().upper() == "EN CONSTRUCCION":
                           estado = "EN CONSTRUCCION"
                           conteo['en_construccion'] += 1
                    elif str(valor_j).strip().upper() == "NO TIENE":
                           estado = "NO TIENE"
                           conteo['no_tiene'] += 1
                    elif str(valor_j).startswith("http"):
                           estado = "LINK VALIDO"
                           conteo['links_validos'] += 1
                    else:
                           estado = "OTRO"
                           conteo['otro'] += 1
                    
                    resultados.append({
                           'fila': row,
                           'nombre': nombre or col_a or f"Fila {row}",
                           'valor_j': valor_j if valor_j else "",
                           'estado': estado
                    })
             
             total = len(resultados)
             sin_link = conteo['en_construccion'] + conteo['no_tiene'] + conteo['vacio'] + conteo['otro']
             
             # Tabla resumen
             ws_resumen['A5'] = "CATEGORIA"
             ws_resumen['B5'] = "CANTIDAD"
             ws_resumen['C5'] = "PORCENTAJE"
             ws_resumen['D5'] = "OBSERVACION"
             for col in ['A5', 'B5', 'C5', 'D5']:
                    ws_resumen[col].font = header_font
                    ws_resumen[col].fill = gris_fill
                    ws_resumen[col].border = borde
             
             datos_resumen = [
                    ("Links Validos (URLs)", conteo['links_validos'], verde_fill, "Docentes con CTI completo"),
                    ("EN CONSTRUCCION", conteo['en_construccion'], amarillo_fill, "CTI en proceso"),
                    ("NO TIENE", conteo['no_tiene'], rojo_fill, "Sin CTI registrado"),
                    ("VACIO (Sin datos)", conteo['vacio'], rojo_fill, "Campo sin completar"),
                    ("OTRO", conteo['otro'], amarillo_fill, "Valor no estandar"),
             ]
             
             for i, (cat, cant, fill, obs) in enumerate(datos_resumen, start=6):
                    ws_resumen[f'A{i}'] = cat
                    ws_resumen[f'B{i}'] = cant
                    ws_resumen[f'C{i}'] = f"{100*cant/total:.1f}%" if total > 0 else "0%"
                    ws_resumen[f'D{i}'] = obs
                    for col in ['A', 'B', 'C', 'D']:
                           ws_resumen[f'{col}{i}'].border = borde
                           ws_resumen[f'{col}{i}'].fill = fill
             
             ws_resumen['A12'] = "TOTAL REGISTROS"
             ws_resumen['B12'] = total
             ws_resumen['A13'] = "CON LINK VALIDO"
             ws_resumen['B13'] = conteo['links_validos']
             ws_resumen['A14'] = "SIN LINK"
             ws_resumen['B14'] = sin_link
             
             ws_resumen.column_dimensions['A'].width = 25
             ws_resumen.column_dimensions['B'].width = 12
             ws_resumen.column_dimensions['C'].width = 12
             ws_resumen.column_dimensions['D'].width = 30
             
             # HOJA 2: DETALLE COMPLETO
             ws_detalle = wb_reporte.create_sheet("DETALLE COMPLETO")
             headers = ["#", "FILA ORIGINAL", "NOMBRE/REFERENCIA", "VALOR COLUMNA J (CTI)", "ESTADO", "TIENE LINK"]
             for col, header in enumerate(headers, start=1):
                    cell = ws_detalle.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = titulo_fill
                    cell.border = borde
             
             for i, r in enumerate(resultados, start=2):
                    ws_detalle.cell(row=i, column=1, value=i-1)
                    ws_detalle.cell(row=i, column=2, value=r['fila'])
                    ws_detalle.cell(row=i, column=3, value=str(r['nombre'])[:50] if r['nombre'] else "")
                    ws_detalle.cell(row=i, column=4, value=r['valor_j'])
                    ws_detalle.cell(row=i, column=5, value=r['estado'])
                    tiene_link = "SI" if r['estado'] == "LINK VALIDO" else "NO"
                    ws_detalle.cell(row=i, column=6, value=tiene_link)
                    
                    if r['estado'] == "LINK VALIDO":
                           fill = verde_fill
                    elif r['estado'] in ["NO TIENE", "VACIO (Sin datos)"]:
                           fill = rojo_fill
                    else:
                           fill = amarillo_fill
                    
                    for col in range(1, 7):
                           ws_detalle.cell(row=i, column=col).fill = fill
                           ws_detalle.cell(row=i, column=col).border = borde
             
             ws_detalle.column_dimensions['A'].width = 5
             ws_detalle.column_dimensions['B'].width = 12
             ws_detalle.column_dimensions['C'].width = 35
             ws_detalle.column_dimensions['D'].width = 60
             ws_detalle.column_dimensions['E'].width = 20
             ws_detalle.column_dimensions['F'].width = 12
             
             # HOJA 3: SIN LINK
             ws_sin_link = wb_reporte.create_sheet("SIN LINK")
             headers_sin = ["#", "FILA ORIGINAL", "NOMBRE/REFERENCIA", "ESTADO", "MOTIVO"]
             for col, header in enumerate(headers_sin, start=1):
                    cell = ws_sin_link.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                    cell.border = borde
             
             idx = 2
             for r in resultados:
                    if r['estado'] != "LINK VALIDO":
                           ws_sin_link.cell(row=idx, column=1, value=idx-1)
                           ws_sin_link.cell(row=idx, column=2, value=r['fila'])
                           ws_sin_link.cell(row=idx, column=3, value=str(r['nombre'])[:50] if r['nombre'] else "")
                           ws_sin_link.cell(row=idx, column=4, value=r['estado'])
                           
                           if r['estado'] == "VACIO (Sin datos)":
                                  motivo = "Campo completamente vacio"
                           elif r['estado'] == "EN CONSTRUCCION":
                                  motivo = "CTI en proceso de creacion"
                           elif r['estado'] == "NO TIENE":
                                  motivo = "Docente indica que no tiene CTI"
                           else:
                                  motivo = "Valor no reconocido"
                           
                           ws_sin_link.cell(row=idx, column=5, value=motivo)
                           
                           for col in range(1, 6):
                                  ws_sin_link.cell(row=idx, column=col).fill = rojo_fill
                                  ws_sin_link.cell(row=idx, column=col).border = borde
                           idx += 1
             
             ws_sin_link.column_dimensions['A'].width = 5
             ws_sin_link.column_dimensions['B'].width = 12
             ws_sin_link.column_dimensions['C'].width = 35
             ws_sin_link.column_dimensions['D'].width = 20
             ws_sin_link.column_dimensions['E'].width = 35
             
             # HOJA 4: CON LINK
             ws_con_link = wb_reporte.create_sheet("CON LINK")
             headers_con = ["#", "FILA ORIGINAL", "NOMBRE/REFERENCIA", "LINK CTI"]
             for col, header in enumerate(headers_con, start=1):
                    cell = ws_con_link.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="006600", end_color="006600", fill_type="solid")
                    cell.border = borde
             
             idx = 2
             for r in resultados:
                    if r['estado'] == "LINK VALIDO":
                           ws_con_link.cell(row=idx, column=1, value=idx-1)
                           ws_con_link.cell(row=idx, column=2, value=r['fila'])
                           ws_con_link.cell(row=idx, column=3, value=str(r['nombre'])[:50] if r['nombre'] else "")
                           ws_con_link.cell(row=idx, column=4, value=r['valor_j'])
                           
                           for col in range(1, 5):
                                  ws_con_link.cell(row=idx, column=col).fill = verde_fill
                                  ws_con_link.cell(row=idx, column=col).border = borde
                           idx += 1
             
             ws_con_link.column_dimensions['A'].width = 5
             ws_con_link.column_dimensions['B'].width = 12
             ws_con_link.column_dimensions['C'].width = 35
             ws_con_link.column_dimensions['D'].width = 70
             
             # Guardar
             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
             filename = f"Analisis_Links_CTI_{timestamp}.xlsx"
             filepath = os.path.join(RESULTADOS_DIR, filename)
             wb_reporte.save(filepath)
             
             return send_from_directory(RESULTADOS_DIR, filename, as_attachment=True)
             
      except Exception as e:
             return jsonify({"error": str(e)}), 500


@app.route('/descargar-excel-detallado')
def descargar_excel_detallado():
      """Genera y descarga Excel detallado con todos los resultados"""
      try:
             resultado = None

             # 1) Preferencia: resultado en memoria (sesión activa)
             if estado_proceso.get("completado") and estado_proceso.get("resultado"):
                    resultado = estado_proceso["resultado"]

             # 2) Fallback: cargar el último JSON guardado en disco
             if not resultado:
                    import json, glob
                    jsons = sorted(
                           glob.glob(os.path.join(RESULTADOS_DIR, 'clasificacion_final_*.json')),
                           key=os.path.getmtime, reverse=True
                    )
                    if jsons:
                           with open(jsons[0], 'r', encoding='utf-8') as _f:
                                  data = json.load(_f)
                           evals_raw = data.get('ranking', [])
                           # Normalizar al formato esperado — incluye TODOS los campos
                           resultado = {
                                  "evaluaciones": [
                                         e if 'puntajes' in e else {
                                                "nombre":               e.get('nombre', ''),
                                                "dni":                  e.get('dni', ''),
                                                "total":                e.get('total', 0),
                                                "maximo":               e.get('maximo', 200),
                                                "porcentaje":           e.get('porcentaje', 0),
                                                "clasificacion":        e.get('clasificacion', ''),
                                                "es_elegible":          e.get('es_elegible', False),
                                                "tipo_perfil":          e.get('tipo_perfil', 'general'),
                                                "puntajes":             e.get('puntajes', {}),
                                                "detalles":             e.get('detalles', {}),
                                                "facultad":             e.get('facultad', ''),
                                                "carrera":              e.get('carrera', ''),
                                                "tipo_candidato":       e.get('tipo_candidato', ''),
                                                "puesto":               e.get('puesto', ''),
                                                "categoria_practitioner": e.get('categoria_practitioner', ''),
                                                "nivel":                e.get('nivel', ''),
                                                "dominio":              e.get('dominio', ''),
                                                "perfil_dominante":     e.get('perfil_dominante', ''),
                                                "perfil_secundario":    e.get('perfil_secundario', ''),
                                                "justificacion_perfil": e.get('justificacion_perfil', ''),
                                                "facultad_recomendada": e.get('facultad_recomendada', ''),
                                                "carrera_recomendada":  e.get('carrera_recomendada', ''),
                                                "cursos_recomendados":  e.get('cursos_recomendados', ''),
                                                "etiqueta_reclutamiento": e.get('etiqueta_reclutamiento', ''),
                                                "url":                  _url_persona(e),
                                                "indice_original":      e.get('indice_original', 0),
                                         } for e in evals_raw
                                  ],
                                  "reportes": {"json": os.path.basename(jsons[0])},
                           }
                           pass  # log removed (encoding issue)

             if not resultado or not resultado.get("evaluaciones"):
                    return jsonify({"error": "No hay resultados disponibles. Ejecuta primero la evaluación."}), 400

             evaluaciones = resultado["evaluaciones"]
             
             # Función auxiliar para obtener valores de detalles de forma segura
             def get_detalle(eval_data, criterio, campo, default=''):
                    detalles = eval_data.get('detalles', {})
                    # Mapear nombres C1, C2, etc. a nombres internos
                    mapeo_nombres = {
                           'C1': 'formacion_academica',
                           'C2': 'experiencia_docente', 
                           'C3': 'experiencia_profesional',
                           'C4': 'centro_labores',
                           'C5': 'produccion_academica',
                           'C6': 'liderazgo_profesional',
                           'C7': 'especializacion'
                    }
                    nombre_interno = mapeo_nombres.get(criterio, criterio)
                    criterio_data = detalles.get(nombre_interno, detalles.get(criterio, {}))
                    if not criterio_data:
                           return default
                    # Buscar el campo o alternativas
                    valor = criterio_data.get(campo)
                    if valor is None:
                           # Campos alternativos
                           if campo == 'categoria':
                                  valor = criterio_data.get('tipo', criterio_data.get('nivel', default))
                           elif campo == 'justificacion':
                                  valor = criterio_data.get('detalle', default)
                    return valor if valor else default
             
             # Columnas reorganizadas por prioridad para RRHH
             datos_excel = []
             for idx, eval in enumerate(evaluaciones):
                    puntajes = eval.get('puntajes', {})
                    det = eval.get('detalles', {})
                    facultad_real = eval.get('facultad', '')
                    carrera_real  = eval.get('carrera', '')
                    # Si el JSON guardado aun dice ARCHIVO SUBIDO, cruzar con 2026.1
                    if facultad_real in ('', 'ARCHIVO SUBIDO'):
                           try:
                                  dm = cargar_datos_requerimientos_para_match()
                                  # DNI: parsear float->int para evitar '167561580' en vez de '16756158'
                                  dni_raw = str(eval.get('dni', '') or '')
                                  try:
                                         dni_k = str(int(float(dni_raw)))
                                  except (ValueError, TypeError):
                                         dni_k = dni_raw.replace('.', '').replace('-', '').upper()
                                  # Nombre: quitar comas para comparar por palabras
                                  nom_palabras = frozenset(
                                         w for w in str(eval.get('nombre', '')).upper().replace(',', ' ').split()
                                         if len(w) > 1
                                  )
                                  nom_k = ' '.join(str(eval.get('nombre', '')).upper().replace(',', ' ').split())
                                  match = (dm['por_dni'].get(dni_k)
                                           or dm['por_nombre'].get(nom_k)
                                           or dm['por_palabras'].get(nom_palabras))
                                  if match:
                                         facultad_real = match.get('facultad', facultad_real)
                                         if not carrera_real or carrera_real in ('SIN CARRERA', 'ARCHIVO SUBIDO'):
                                                carrera_real = match.get('carrera', carrera_real)
                           except Exception:
                                  pass

                    datos_excel.append({
                           '#':             idx + 1,
                           'DNI':           eval.get('dni', ''),
                           'Nombre':        eval.get('nombre', 'Sin nombre'),
                           'Facultad':      facultad_real or 'Sin asignar',
                           'Carrera':       carrera_real  or 'Sin asignar',
                           'Tipo Candidato': eval.get('tipo_candidato', ''),
                           'Puesto':        eval.get('puesto', ''),
                           'Puntaje':       eval.get('total', 0),
                           'Maximo':        eval.get('maximo', 200),
                           'Elegible':      'SI' if eval.get('es_elegible', False) else 'NO',
                           'Clasificacion': eval.get('clasificacion', '').split(' (')[0],
                           'Rol Dominante': eval.get('perfil_dominante', ''),
                           'Dominio':       eval.get('dominio', ''),
                           'Etiqueta':      eval.get('etiqueta_reclutamiento', ''),
                           'Justif. Talento': eval.get('justificacion_perfil', ''),
                           'C1 Formacion':  f"{puntajes.get('C1', 0)}/{det.get('formacion_academica', {}).get('maximo', 50)}",
                           'Justif. C1':    get_detalle(eval, 'C1', 'justificacion', ''),
                           'C2 Docencia':   f"{puntajes.get('C2', 0)}/{det.get('experiencia_docente', {}).get('maximo', 40)}",
                           'Justif. C2':    get_detalle(eval, 'C2', 'justificacion', ''),
                           'C3 Experiencia': f"{puntajes.get('C3', 0)}/{det.get('experiencia_profesional', {}).get('maximo', 40)}",
                           'Justif. C3':    get_detalle(eval, 'C3', 'justificacion', ''),
                           'C4 Centro Lab.': f"{puntajes.get('C4', 0)}/{det.get('centro_labores', {}).get('maximo', 20)}",
                           'C5 Produccion': f"{puntajes.get('C5', 0)}/{det.get('produccion_academica', {}).get('maximo', 50)}",
                           'Justif. C5':    get_detalle(eval, 'C5', 'justificacion', ''),
                           'URL CTI':       _url_persona(eval),
                    })
             
             df = pd.DataFrame(datos_excel)
             
             # Generar archivo
             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
             filename = f"Ranking_Docentes_{timestamp}.xlsx"
             filepath = os.path.join(RESULTADOS_DIR, filename)
             
             # â" Colores por Nivel â"
             NIVEL_COLORES = {
                    "A+": "1E3A8A",   # Azul oscuro
                    "A":  "1D4ED8",   # Azul medio
                    "B":  "2563EB",   # Azul claro
                    "C":  "64748B",   # Gris azul
                    "D":  "94A3B8",   # Gris claro
                    "E":  "CBD5E1",   # Gris muy claro
             }
             NIVEL_COLOR_TEXTO = {
                    "A+": "FFFFFF", "A": "FFFFFF", "B": "FFFFFF",
                    "C": "FFFFFF", "D": "1E293B", "E": "1E293B",
             }
             ELEGIBLE_COLORES = {"SI": "D1FAE5", "NO": "FEE2E2"}
             
             with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Ranking Docentes', index=False)
                    ws = writer.sheets['Ranking Docentes']
                    
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
                    
                    BORDE = Border(
                           left=Side(style='thin', color='CBD5E1'),
                           right=Side(style='thin', color='CBD5E1'),
                           top=Side(style='thin', color='CBD5E1'),
                           bottom=Side(style='thin', color='CBD5E1'),
                    )
                    HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
                    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                    
                    # Columnas con justificacion (texto largo, wrap)
                    COLS_JUSTIF = {'Justif. Talento', 'Justif. C1', 'Justif. C2',
                                   'Justif. C3', 'Justif. C5', 'Etiqueta'}
                    COLS_URL    = {'URL CTI'}

                    # Ancho de columnas (en unidades de openpyxl ~= caracteres)
                    ANCHOS = {
                           '#': 5, 'DNI': 13, 'Nombre': 35,
                           'Facultad': 20, 'Carrera': 22,
                           'Tipo Candidato': 18, 'Puesto': 16,
                           'Puntaje': 10, 'Maximo': 10, 'Elegible': 10,
                           'Clasificacion': 20, 'Rol Dominante': 24, 'Dominio': 18,
                           'Etiqueta': 50,
                           'Justif. Talento': 60,
                           'C1 Formacion': 14, 'Justif. C1': 55,
                           'C2 Docencia':  14, 'Justif. C2': 55,
                           'C3 Experiencia': 14, 'Justif. C3': 55,
                           'C4 Centro Lab.': 14,
                           'C5 Produccion': 14, 'Justif. C5': 55,
                           'URL CTI': 50,
                    }

                    # Encabezados
                    for col_idx, cell in enumerate(ws[1], 1):
                           header_name = str(cell.value)
                           cell.fill = HEADER_FILL
                           cell.font = HEADER_FONT
                           cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                           cell.border = BORDE
                           ws.row_dimensions[1].height = 36
                           col_letter = cell.column_letter
                           ws.column_dimensions[col_letter].width = ANCHOS.get(header_name, 18)

                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = ws.dimensions

                    # Filas de datos
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                           # Altura fija para que se vea el texto largo sin tener que expandir
                           ws.row_dimensions[row_idx].height = 90

                           for col_idx, cell in enumerate(row, 1):
                                  header = ws.cell(row=1, column=col_idx).value
                                  cell.border = BORDE

                                  if header in COLS_JUSTIF:
                                         cell.alignment = Alignment(vertical="top", wrap_text=True)
                                         cell.font = Font(size=9, name="Calibri")
                                  elif header in COLS_URL:
                                         cell.alignment = Alignment(vertical="top", wrap_text=False)
                                         if cell.value:
                                                cell.font = Font(color="1D4ED8", underline="single", size=9, name="Calibri")
                                  elif header in ('C1 Formacion', 'C2 Docencia', 'C3 Experiencia',
                                                  'C4 Centro Lab.', 'C5 Produccion', 'Puntaje', 'Maximo'):
                                         cell.alignment = Alignment(horizontal="center", vertical="top")
                                         cell.font = Font(bold=True, size=9, name="Calibri")
                                  else:
                                         cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                                         cell.font = Font(size=9, name="Calibri")


                                  # Colorear columna Elegible
                                  if header == 'Elegible' and str(cell.value).upper() in ELEGIBLE_COLORES:
                                         color = ELEGIBLE_COLORES[str(cell.value).upper()]
                                         cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                         cell.font = Font(bold=True, size=9, name="Calibri")

                    # Filas alternas (gris muy claro)
                    for row_idx in range(2, ws.max_row + 1):
                           if row_idx % 2 == 0:
                                  for cell in ws[row_idx]:
                                        if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                                               cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
             
             return send_from_directory(RESULTADOS_DIR, filename, as_attachment=True)

             
      except Exception as e:
             return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
      import sys, io
      # Asegurar UTF-8 en la consola de Windows
      if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
             sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
      print("\n" + "="*70)
      print(">>> PEOPLE ANALYTICS - TALENTO & CULTURA USIL")
      print("="*70)
      print("\nServidor iniciado: http://localhost:5000")
      print("Sistema de analisis inteligente de perfiles docentes")
      print("\nPresiona Ctrl+C para detener\n")
      
      app.run(host='127.0.0.1', debug=False, port=5000, use_reloader=False, threaded=True)

