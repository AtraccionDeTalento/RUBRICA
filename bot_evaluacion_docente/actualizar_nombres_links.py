"""
Script para agregar nombres a la hoja de links
Lee los nombres del Excel de requerimientos y los agrega a la columna B
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import sys

# Rutas de archivos
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LINKS_PATH = os.path.join(BASE_DIR, "LINKS", "HOJA DE PRUEBA de Extraccion de LINKs.xlsx")
REQUERIMIENTOS_PATH = os.path.join(BASE_DIR, "Extraccion de data", "Requerimiento docentes 2026-1 300126.xlsx")


def obtener_mapeo_url_nombre():
    """Lee el Excel de requerimientos y crea un diccionario URL -> Nombre"""
    print("Leyendo archivo de requerimientos...")
    df_req = pd.read_excel(REQUERIMIENTOS_PATH, sheet_name='2026.1')
    print(f"  Total registros: {len(df_req)}")
    
    # Crear diccionario de URL -> Nombre
    url_to_nombre = {}
    url_to_dni = {}
    
    for idx, row in df_req.iterrows():
        url = str(row.get('CTI', '')).strip()
        nombre = str(row.get('APELLIDOS Y NOMBRES DEL CANDIDATO', '')).strip()
        dni = str(row.get('DNI', '')).strip()
        
        if url.startswith('http') and 'ctivitae.concytec.gob.pe' in url:
            if nombre and nombre.lower() not in ['nan', 'none', '']:
                url_to_nombre[url] = nombre
            if dni and dni.lower() not in ['nan', 'none', '']:
                url_to_dni[url] = dni
    
    print(f"  URLs con nombre mapeado: {len(url_to_nombre)}")
    print(f"  URLs con DNI mapeado: {len(url_to_dni)}")
    
    return url_to_nombre, url_to_dni


def actualizar_excel_links():
    """Actualiza el Excel de links agregando nombres en columna B y DNI en columna C"""
    
    # Verificar que los archivos existen
    if not os.path.exists(REQUERIMIENTOS_PATH):
        print(f"ERROR: No se encontro el archivo de requerimientos:\n  {REQUERIMIENTOS_PATH}")
        return False
    
    if not os.path.exists(LINKS_PATH):
        print(f"ERROR: No se encontro el archivo de links:\n  {LINKS_PATH}")
        return False
    
    # Obtener mapeo URL -> Nombre
    url_to_nombre, url_to_dni = obtener_mapeo_url_nombre()
    
    print(f"\nAbriendo archivo de links: {os.path.basename(LINKS_PATH)}")
    
    try:
        wb = load_workbook(LINKS_PATH)
        ws = wb.active
        
        # Obtener el nombre actual de la columna A (fila 1)
        titulo_a = ws.cell(row=1, column=1).value or "CTI (Links)"
        
        # Configurar encabezados
        ws.cell(row=1, column=1, value=titulo_a)
        ws.cell(row=1, column=2, value="NOMBRE CANDIDATO")
        ws.cell(row=1, column=3, value="DNI")
        
        # Estilo para encabezados
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        
        for col in [1, 2, 3]:
            cell = ws.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 15
        
        # Procesar cada fila
        total_actualizados = 0
        total_sin_nombre = 0
        
        for row_num in range(2, ws.max_row + 1):
            url = ws.cell(row=row_num, column=1).value
            
            if url and isinstance(url, str) and url.startswith('http'):
                url_clean = url.strip()
                
                # Buscar nombre
                nombre = url_to_nombre.get(url_clean, "")
                dni = url_to_dni.get(url_clean, "")
                
                if nombre:
                    ws.cell(row=row_num, column=2, value=nombre)
                    total_actualizados += 1
                else:
                    ws.cell(row=row_num, column=2, value="(Sin nombre en Excel)")
                    total_sin_nombre += 1
                
                if dni:
                    ws.cell(row=row_num, column=3, value=dni)
            
            elif url and isinstance(url, str):
                # Es texto pero no URL (ej: "EN CONSTRUCCION")
                ws.cell(row=row_num, column=2, value="N/A")
                ws.cell(row=row_num, column=3, value="N/A")
        
        # Guardar
        print("\nGuardando cambios...")
        wb.save(LINKS_PATH)
        wb.close()
        
        print(f"\n{'='*60}")
        print("ACTUALIZACION COMPLETADA")
        print(f"{'='*60}")
        print(f"  Links con nombre agregado: {total_actualizados}")
        print(f"  Links sin nombre encontrado: {total_sin_nombre}")
        print(f"\nArchivo actualizado: {LINKS_PATH}")
        
        return True
        
    except PermissionError:
        print("\n" + "="*60)
        print("ERROR: El archivo Excel esta ABIERTO")
        print("="*60)
        print("Por favor CIERRA el archivo Excel e intenta nuevamente")
        print(f"Archivo: {os.path.basename(LINKS_PATH)}")
        print("="*60)
        return False
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    print("="*60)
    print("ACTUALIZADOR DE NOMBRES EN HOJA DE LINKS")
    print("="*60)
    print()
    
    exito = actualizar_excel_links()
    
    if exito:
        print("\n[OK] Proceso completado exitosamente")
    else:
        print("\n[ERROR] El proceso no se completo")
    
    sys.exit(0 if exito else 1)
