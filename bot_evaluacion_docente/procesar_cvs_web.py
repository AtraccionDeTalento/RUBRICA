"""
Script principal para procesar CVs desde enlaces web
Lee archivo Excel con URLs y ejecuta el sistema de clasificación
"""
import os
import sys
import pandas as pd
from extractor_web_cvs import ExtractorWebCVs
from motor_evaluacion import MotorEvaluacion
from generador_decisiones import GeneradorDecisiones
from generador_reportes import GeneradorReportes
from config import RESULTADOS_DIR

# Configurar encoding para Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass

# Ruta del archivo Excel con los links
ARCHIVO_LINKS = r"c:\Users\jlopezp\OneDrive - Universidad San Ignacio de Loyola\PROYECTO RUBRICA\LINKS\HOJA DE PRUEBA de Extraccion de LINKs.xlsx"


def leer_links_desde_excel(ruta_excel: str):
    """
    Lee los links desde el archivo Excel
    
    Args:
        ruta_excel: Ruta del archivo Excel
        
    Returns:
        Lista de URLs válidas
    """
    try:
        print(f"\n📄 Leyendo archivo: {os.path.basename(ruta_excel)}")
        
        # Intentar con openpyxl (más robusto con archivos bloqueados)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(ruta_excel, read_only=True, data_only=True)
            ws = wb.active
            
            urls = []
            for row in ws.iter_rows(values_only=True):
                for valor in row:
                    if valor and isinstance(valor, str):
                        if valor.startswith('http://') or valor.startswith('https://'):
                            urls.append(valor.strip())
            
            wb.close()
            
            print(f"   ✓ Se encontraron {len(urls)} enlaces")
            return urls
            
        except Exception as e1:
            # Si falla openpyxl, intentar con pandas
            print(f"   ⚠️  Intentando método alternativo...")
            df = pd.read_excel(ruta_excel, header=None)
            
            # Extraer URLs (buscar celdas que contengan http/https)
            urls = []
            
            for col in df.columns:
                for valor in df[col]:
                    if pd.notna(valor) and isinstance(valor, str):
                        if valor.startswith('http://') or valor.startswith('https://'):
                            urls.append(valor.strip())
            
            print(f"   ✓ Se encontraron {len(urls)} enlaces")
            
            return urls
        
    except Exception as e:
        print(f"   ❌ Error al leer archivo: {e}")
        print(f"\n💡 Consejo: Cierra el archivo Excel si está abierto e intenta de nuevo")
        return []


def main():
    """Función principal"""
    
    print("="*70)
    print("🌐 SISTEMA DE EVALUACIÓN DOCENTE - EXTRACCIÓN WEB")
    print("="*70)
    
    # PASO 1: Leer links del Excel
    print("\nOpciones de entrada:")
    print("  1. Leer desde archivo Excel")
    print("  2. Ingresar URLs manualmente (separadas por enter)")
    print("  3. Usar URLs de prueba")
    
    opcion = input("\nSelecciona una opción (1-3): ").strip()
    
    urls = []
    
    if opcion == "1":
        urls = leer_links_desde_excel(ARCHIVO_LINKS)
    elif opcion == "2":
        print("\nIngresa las URLs (una por línea, línea vacía para terminar):")
        while True:
            url = input().strip()
            if not url:
                break
            if url.startswith('http://') or url.startswith('https://'):
                urls.append(url)
            else:
                print("  ⚠️  URL inválida, debe empezar con http:// o https://")
    elif opcion == "3":
        # URLs de prueba
        urls = [
            "https://ctivitae.concytec.gob.pe/appDirectorioCTI/VerDatosInvestigador.do?id_investigador=440291",
            "https://ctivitae.concytec.gob.pe/appDirectorioCTI/VerDatosInvestigador.do?id_investigador=472410"
        ]
        print(f"\n✓ Usando {len(urls)} URLs de prueba")
    else:
        print("\n❌ Opción inválida")
        return
    
    if not urls:
        print("\n❌ No se encontraron enlaces válidos")
        return
    
    print(f"\n📋 Enlaces a procesar:")
    for i, url in enumerate(urls, 1):
        print(f"   {i}. {url}")
    
    # Confirmar procesamiento (solo si es interactivo)
    print(f"\n⚠️  Se procesarán {len(urls)} perfiles de CTI Vitae")
    
    try:
        respuesta = input("¿Desea continuar? (s/n): ").strip().lower()
        if respuesta != 's':
            print("\n❌ Proceso cancelado")
            return
    except EOFError:
        # Si no hay entrada (pipe o batch), continuar automáticamente
        print("  → Continuando automáticamente...")
        pass
    
    # PASO 2: Extraer información de las webs
    extractor = ExtractorWebCVs()
    cvs_extraidos = extractor.extraer_multiples_cvs(urls)
    
    # PASO 3: Evaluar con el motor de evaluación
    print("\n" + "="*70)
    print("EVALUANDO CANDIDATOS CON RÚBRICA USIL")
    print("="*70)
    
    # Importar analizador de rúbrica
    from analizador_rubrica import analizar_rubrica_completa
    
    # Analizar rúbrica
    rubrica = analizar_rubrica_completa()
    
    evaluaciones = []
    motor = MotorEvaluacion(
        criterios=rubrica['estructura']['criterios'],
        pesos=rubrica['pesos']
    )
    
    for i, cv_data in enumerate(cvs_extraidos, 1):
        print(f"\n👤 [{i}/{len(cvs_extraidos)}] Evaluando: {cv_data['nombre']}")
        
        try:
            evaluacion = motor.evaluar_cv_completo(cv_data)
            evaluacion['archivo'] = cv_data.get('url', 'URL no disponible')
            evaluacion['datos_cv'] = cv_data
            
            evaluaciones.append(evaluacion)
            
            print(f"   ✅ Puntuación total: {evaluacion['puntuacion_total']:.0f}/230")
            
        except Exception as e:
            print(f"   ❌ Error al evaluar: {e}")
    
    # PASO 4: Generar clasificación
    print("\n" + "="*70)
    print("GENERANDO CLASIFICACIÓN DE PERFILES")
    print("="*70)
    
    generador_decisiones = GeneradorDecisiones(evaluaciones)
    clasificacion = generador_decisiones.generar_clasificacion_completa()
    
    # Mostrar tabla consolidada
    tabla_consolidada = generador_decisiones.generar_tabla_consolidada()
    print(tabla_consolidada)
    
    # PASO 5: Generar reportes
    print("\n" + "="*70)
    print("GENERANDO REPORTES")
    print("="*70)
    
    generador_reportes = GeneradorReportes(evaluaciones, clasificacion)
    
    ruta_excel = generador_reportes.generar_excel_comparativo()
    ruta_json = generador_reportes.generar_json_decision()
    
    # RESUMEN FINAL
    print("\n" + "="*70)
    print("✅ PROCESO COMPLETADO")
    print("="*70)
    
    print(f"\n📊 Resumen por perfil:")
    for perfil, candidatos in clasificacion['resumen_por_perfil'].items():
        print(f"   • {perfil}: {len(candidatos)} candidato(s)")
        for nombre in candidatos:
            print(f"      - {nombre}")
    
    print(f"\n📥 Reportes generados:")
    print(f"   • Excel: {os.path.basename(ruta_excel)}")
    print(f"   • JSON: {os.path.basename(ruta_json)}")
    print(f"\n📂 Ubicación: {RESULTADOS_DIR}")
    
    print("\n" + "="*70)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n❌ Proceso interrumpido por el usuario")
        sys.exit(1)
    except Exception as e:
        print(f"\n\n❌ Error crítico: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
